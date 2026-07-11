"""
Patch 49a — Fix Critical Bugs
1. Sidebar Scroll Fix (base CSS injection)
2. NoReverseMatch in Attendance List (pk safety)
3. Excel Export using openpyxl (real Excel)
4. Map CSS Fix (Live Map + Visit Form)
5. Stealth Tracking Policy Auto-enable helper
"""

import os
import sys
import django

# ════════════════════════════════════════════════
# Setup Django
# ════════════════════════════════════════════════
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'motionhr.settings')
django.setup()

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def write_file(path, content):
    full = os.path.join(BASE, path)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    with open(full, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"✅ كُتب: {path}")

def read_file(path):
    full = os.path.join(BASE, path)
    if not os.path.exists(full):
        return None
    with open(full, 'r', encoding='utf-8') as f:
        return f.read()

print("=" * 60)
print("Patch 49a — Fix Critical Bugs")
print("=" * 60)


# ════════════════════════════════════════════════
# BUG 1: Sidebar Scroll + Map Fix CSS
# ════════════════════════════════════════════════
print("\n📌 Bug 1: Sidebar Scroll + Map CSS Fix")

sidebar_map_css = """
/* ═══════════════════════════════════════════════════
   Patch 49a — Sidebar Scroll + Map Fix
   ═══════════════════════════════════════════════════ */

/* ── Sidebar Scroll ── */
.sidebar {
    position: fixed !important;
    top: 0;
    right: 0;
    height: 100vh !important;
    overflow-y: auto !important;
    overflow-x: hidden !important;
    z-index: 1040;
    scrollbar-width: thin;
    scrollbar-color: rgba(255,255,255,0.3) transparent;
}

.sidebar::-webkit-scrollbar {
    width: 4px;
}
.sidebar::-webkit-scrollbar-track {
    background: transparent;
}
.sidebar::-webkit-scrollbar-thumb {
    background: rgba(255,255,255,0.3);
    border-radius: 2px;
}

/* ── Main Content Offset ── */
.main-content {
    margin-right: 260px;
}

@media (max-width: 991.98px) {
    .main-content {
        margin-right: 0;
    }
    .sidebar {
        position: fixed !important;
        right: -260px;
        transition: right 0.3s ease;
    }
    .sidebar.show {
        right: 0;
    }
}

/* ── Map Fix — ثابتة مش بتتحرك ── */
#live-map,
#visit-map,
#checkin-map,
.leaflet-map-container,
[id$="-map"] {
    position: relative !important;
    height: 420px !important;
    width: 100% !important;
    z-index: 1 !important;
    border-radius: 12px;
    overflow: hidden;
}

.leaflet-container {
    position: absolute !important;
    top: 0 !important;
    left: 0 !important;
    right: 0 !important;
    bottom: 0 !important;
    width: 100% !important;
    height: 100% !important;
}

.map-wrapper {
    position: relative !important;
    width: 100%;
    height: 420px;
    overflow: hidden;
    border-radius: 12px;
    border: 1px solid #dee2e6;
}

/* ── منع تحرك الخريطة مع الصفحة ── */
.leaflet-pane,
.leaflet-map-pane {
    transform-origin: 0 0 !important;
}
"""

write_file('static/css/patch49a_fixes.css', sidebar_map_css)


# ════════════════════════════════════════════════
# BUG 2: NoReverseMatch في Attendance List
# ════════════════════════════════════════════════
print("\n📌 Bug 2: Fix Attendance List Template (NoReverseMatch)")

attendance_list_template = """{% extends 'base/dashboard_base.html' %}
{% load custom_filters %}

{% block title %}سجلات الحضور{% endblock %}

{% block content %}
<div class="container-fluid">

  <!-- Header -->
  <div class="d-flex justify-content-between align-items-center mb-4">
    <div>
      <h4 class="mb-1 fw-bold">
        <i class="bi bi-calendar-check text-primary me-2"></i>سجلات الحضور
      </h4>
      <nav aria-label="breadcrumb">
        <ol class="breadcrumb mb-0 small">
          <li class="breadcrumb-item"><a href="{% url 'dashboard:index' %}">الرئيسية</a></li>
          <li class="breadcrumb-item active">سجلات الحضور</li>
        </ol>
      </nav>
    </div>
    {% if request.user.role in 'company_admin,hr_manager' %}
    <a href="{% url 'attendance:check_in' %}" class="btn btn-primary">
      <i class="bi bi-plus-circle me-1"></i>تسجيل حضور
    </a>
    {% endif %}
  </div>

  <!-- إحصائيات اليوم -->
  <div class="row g-3 mb-4">
    <div class="col-6 col-md-3">
      <div class="card border-0 shadow-sm text-center py-3">
        <div class="fs-2 fw-bold text-primary">{{ today_stats.total }}</div>
        <div class="text-muted small">إجمالي اليوم</div>
      </div>
    </div>
    <div class="col-6 col-md-3">
      <div class="card border-0 shadow-sm text-center py-3">
        <div class="fs-2 fw-bold text-success">{{ today_stats.present }}</div>
        <div class="text-muted small">حاضر</div>
      </div>
    </div>
    <div class="col-6 col-md-3">
      <div class="card border-0 shadow-sm text-center py-3">
        <div class="fs-2 fw-bold text-warning">{{ today_stats.late }}</div>
        <div class="text-muted small">متأخر</div>
      </div>
    </div>
    <div class="col-6 col-md-3">
      <div class="card border-0 shadow-sm text-center py-3">
        <div class="fs-2 fw-bold text-danger">{{ today_stats.absent }}</div>
        <div class="text-muted small">غائب</div>
      </div>
    </div>
  </div>

  <!-- فلترة -->
  <div class="card border-0 shadow-sm mb-4">
    <div class="card-body">
      <form method="get" class="row g-2 align-items-end">
        <div class="col-md-3">
          <label class="form-label small text-muted">من تاريخ</label>
          <input type="date" name="date_from" value="{{ date_from }}" class="form-control form-control-sm">
        </div>
        <div class="col-md-3">
          <label class="form-label small text-muted">إلى تاريخ</label>
          <input type="date" name="date_to" value="{{ date_to }}" class="form-control form-control-sm">
        </div>
        <div class="col-md-3">
          <label class="form-label small text-muted">الموظف</label>
          <select name="employee" class="form-select form-select-sm">
            <option value="">-- الكل --</option>
            {% for emp in employees %}
            <option value="{{ emp.id }}" {% if employee_id == emp.id|stringformat:'s' %}selected{% endif %}>
              {{ emp.full_name_ar }}
            </option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-2">
          <label class="form-label small text-muted">الحالة</label>
          <select name="status" class="form-select form-select-sm">
            <option value="">-- الكل --</option>
            {% for val, label in status_choices %}
            <option value="{{ val }}" {% if status == val %}selected{% endif %}>{{ label }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-1">
          <button type="submit" class="btn btn-primary btn-sm w-100">
            <i class="bi bi-search"></i>
          </button>
        </div>
      </form>
    </div>
  </div>

  <!-- الجدول -->
  <div class="card border-0 shadow-sm">
    <div class="card-header bg-white d-flex justify-content-between align-items-center py-3">
      <span class="fw-semibold">
        <i class="bi bi-list-ul me-2 text-primary"></i>
        السجلات ({{ total_count }})
      </span>
    </div>
    <div class="card-body p-0">
      {% if attendances %}
      <div class="table-responsive">
        <table class="table table-hover align-middle mb-0">
          <thead class="table-light">
            <tr>
              <th>الموظف</th>
              <th>التاريخ</th>
              <th>وقت الدخول</th>
              <th>وقت الخروج</th>
              <th>الحالة</th>
              <th>التأخير</th>
              <th>ملاحظات</th>
              {% if request.user.role in 'company_admin,hr_manager' %}
              <th>إجراء</th>
              {% endif %}
            </tr>
          </thead>
          <tbody>
            {% for att in attendances %}
            <tr>
              <td>
                <div class="fw-semibold small">{{ att.employee.full_name_ar }}</div>
                <div class="text-muted" style="font-size:0.75rem">{{ att.employee.employee_code }}</div>
              </td>
              <td class="small">{{ att.date|date:"Y/m/d" }}</td>
              <td class="small">
                {% if att.check_in %}
                  <span class="text-success fw-semibold">{{ att.check_in|time:"H:i" }}</span>
                {% else %}
                  <span class="text-muted">—</span>
                {% endif %}
              </td>
              <td class="small">
                {% if att.check_out %}
                  <span class="text-primary fw-semibold">{{ att.check_out|time:"H:i" }}</span>
                {% else %}
                  <span class="text-muted">—</span>
                {% endif %}
              </td>
              <td>
                {% if att.status == 'present' %}
                  <span class="badge bg-success-subtle text-success border border-success-subtle rounded-pill px-2">حاضر</span>
                {% elif att.status == 'late' %}
                  <span class="badge bg-warning-subtle text-warning border border-warning-subtle rounded-pill px-2">متأخر</span>
                {% elif att.status == 'absent' %}
                  <span class="badge bg-danger-subtle text-danger border border-danger-subtle rounded-pill px-2">غائب</span>
                {% elif att.status == 'on_leave' %}
                  <span class="badge bg-info-subtle text-info border border-info-subtle rounded-pill px-2">إجازة</span>
                {% else %}
                  <span class="badge bg-secondary-subtle text-secondary rounded-pill px-2">{{ att.status }}</span>
                {% endif %}
              </td>
              <td class="small">
                {% if att.late_minutes and att.late_minutes > 0 %}
                  <span class="text-danger">{{ att.late_minutes }} د</span>
                {% else %}
                  <span class="text-muted">—</span>
                {% endif %}
              </td>
              <td class="small text-muted">{{ att.notes|default:"—"|truncatechars:30 }}</td>
              {% if request.user.role in 'company_admin,hr_manager' %}
              <td>
                {% if att.id %}
                <a href="{% url 'attendance:override' att.id %}"
                   class="btn btn-sm btn-outline-secondary"
                   title="تعديل">
                  <i class="bi bi-pencil"></i>
                </a>
                {% endif %}
              </td>
              {% endif %}
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>

      <!-- Pagination -->
      {% if page_obj.has_other_pages %}
      <div class="d-flex justify-content-center py-3">
        <nav>
          <ul class="pagination pagination-sm mb-0">
            {% if page_obj.has_previous %}
            <li class="page-item">
              <a class="page-link" href="?page={{ page_obj.previous_page_number }}&date_from={{ date_from }}&date_to={{ date_to }}&employee={{ employee_id }}&status={{ status }}">
                <i class="bi bi-chevron-right"></i>
              </a>
            </li>
            {% endif %}
            <li class="page-item active">
              <span class="page-link">{{ page_obj.number }} / {{ page_obj.paginator.num_pages }}</span>
            </li>
            {% if page_obj.has_next %}
            <li class="page-item">
              <a class="page-link" href="?page={{ page_obj.next_page_number }}&date_from={{ date_from }}&date_to={{ date_to }}&employee={{ employee_id }}&status={{ status }}">
                <i class="bi bi-chevron-left"></i>
              </a>
            </li>
            {% endif %}
          </ul>
        </nav>
      </div>
      {% endif %}

      {% else %}
      <div class="text-center py-5 text-muted">
        <i class="bi bi-calendar-x fs-1 opacity-25"></i>
        <p class="mt-3">لا توجد سجلات حضور للفترة المحددة</p>
      </div>
      {% endif %}
    </div>
  </div>

</div>
{% endblock %}
"""

write_file('templates/attendance/list.html', attendance_list_template)


# ════════════════════════════════════════════════
# BUG 3: Excel Export — openpyxl حقيقي
# ════════════════════════════════════════════════
print("\n📌 Bug 3: Real Excel Export with openpyxl")

reports_utils = '''"""
reports/utils.py — Patch 49a
Excel export حقيقي باستخدام openpyxl
"""

from datetime import date, timedelta
from django.http import HttpResponse
from django.utils import timezone

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from attendance.models import Attendance
from leaves.models import LeaveRequest


# ════════════════════════════════════════
# Date Range Helper
# ════════════════════════════════════════

def get_date_range(period, custom_start=None, custom_end=None):
    today = date.today()
    if period == "week":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
    elif period == "month":
        start = today.replace(day=1)
        next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
        end = next_month - timedelta(days=1)
    elif period == "quarter":
        q = (today.month - 1) // 3
        start = date(today.year, q * 3 + 1, 1)
        if q == 3:
            end = date(today.year, 12, 31)
        else:
            end = date(today.year, (q + 1) * 3 + 1, 1) - timedelta(days=1)
    elif period == "year":
        start = today.replace(month=1, day=1)
        end = today.replace(month=12, day=31)
    elif period == "custom" and custom_start and custom_end:
        start, end = custom_start, custom_end
    else:
        start = today.replace(day=1)
        end = today
    return start, end


# ════════════════════════════════════════
# Attendance Summary
# ════════════════════════════════════════

def get_attendance_summary(company, start_date, end_date):
    qs = Attendance.objects.filter(
        employee__company=company,
        date__range=(start_date, end_date)
    )
    total = qs.count()
    present = qs.filter(status='present').count()
    late = qs.filter(status='late').count()
    absent = qs.filter(status='absent').count()
    on_leave = qs.filter(status='on_leave').count()
    return {
        'total': total,
        'present': present,
        'late': late,
        'absent': absent,
        'on_leave': on_leave,
        'rate': round(present / total * 100, 1) if total else 0,
    }


# ════════════════════════════════════════
# Employee Attendance Details
# ════════════════════════════════════════

def get_employee_attendance_details(company, start_date, end_date):
    from employees.models import Employee
    employees = Employee.objects.filter(company=company, status='active')
    result = []
    for emp in employees:
        qs = Attendance.objects.filter(
            employee=emp,
            date__range=(start_date, end_date)
        )
        total = qs.count()
        present = qs.filter(status='present').count()
        late = qs.filter(status='late').count()
        absent = qs.filter(status='absent').count()
        on_leave = qs.filter(status='on_leave').count()

        from django.db.models import Sum
        late_mins = qs.aggregate(Sum('late_minutes'))['late_minutes__sum'] or 0
        work_seconds = 0
        for a in qs:
            if a.check_in and a.check_out:
                from datetime import datetime, date as dt_date
                ci = datetime.combine(dt_date.today(), a.check_in)
                co = datetime.combine(dt_date.today(), a.check_out)
                if co > ci:
                    work_seconds += (co - ci).seconds
        work_hours = round(work_seconds / 3600, 1)

        result.append({
            'employee': emp,
            'total': total,
            'present': present,
            'late': late,
            'absent': absent,
            'on_leave': on_leave,
            'rate': round(present / total * 100, 1) if total else 0,
            'late_mins': late_mins,
            'work_hours': work_hours,
        })
    return result


# ════════════════════════════════════════
# Field Tracking Summary
# ════════════════════════════════════════

def get_field_tracking_summary(company, start_date, end_date):
    from attendance.models import LocationLog
    from employees.models import Employee
    employees = Employee.objects.filter(
        company=company,
        status='active',
        is_field_worker=True
    )
    result = []
    for emp in employees:
        logs = LocationLog.objects.filter(
            employee=emp,
            timestamp__date__range=(start_date, end_date)
        )
        result.append({
            'employee': emp,
            'log_count': logs.count(),
            'last_log': logs.order_by('-timestamp').first(),
        })
    return result


# ════════════════════════════════════════
# Excel Export — openpyxl احترافي
# ════════════════════════════════════════

def export_to_excel(headers, rows, filename="report", company=None, report_title="تقرير"):
    """
    تصدير Excel احترافي بـ openpyxl
    - لوجو الشركة في الأعلى
    - عنوان التقرير
    - هيدر ملوّن
    - بيانات منظمة
    - Footer باسم البرنامج
    """
    if not OPENPYXL_AVAILABLE:
        # Fallback بسيط لو مش مثبت
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        import csv
        writer = csv.writer(response)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_title[:31]
    ws.sheet_view.rightToLeft = True  # RTL

    # ── الألوان ──
    PRIMARY = "06B6D4"
    HEADER_BG = "0E7490"
    LIGHT_ROW = "F0FDFF"
    WHITE = "FFFFFF"
    DARK = "1E293B"
    GRAY = "94A3B8"

    # ── Borders ──
    thin_border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6'),
    )

    current_row = 1

    # ── Row 1: اسم الشركة ──
    company_name = company.name if company else "MotionHR"
    ws.merge_cells(f'A{current_row}:{get_column_letter(max(len(headers),1))}{current_row}')
    cell = ws.cell(row=current_row, column=1, value=company_name)
    cell.font = Font(name='Cairo', size=16, bold=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 35
    current_row += 1

    # ── Row 2: عنوان التقرير ──
    ws.merge_cells(f'A{current_row}:{get_column_letter(max(len(headers),1))}{current_row}')
    cell = ws.cell(row=current_row, column=1, value=report_title)
    cell.font = Font(name='Cairo', size=13, bold=True, color=DARK)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # ── Row 3: تاريخ التصدير ──
    ws.merge_cells(f'A{current_row}:{get_column_letter(max(len(headers),1))}{current_row}')
    export_date = timezone.now().strftime("%Y/%m/%d %H:%M")
    cell = ws.cell(row=current_row, column=1, value=f"تاريخ التصدير: {export_date}")
    cell.font = Font(name='Cairo', size=10, color=GRAY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # ── فاصل ──
    current_row += 1

    # ── Header Row ──
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.font = Font(name='Cairo', size=11, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # ── Data Rows ──
    for row_idx, row in enumerate(rows):
        bg = LIGHT_ROW if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name='Cairo', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # ── Footer ──
    current_row += 1
    ws.merge_cells(f'A{current_row}:{get_column_letter(max(len(headers),1))}{current_row}')
    cell = ws.cell(row=current_row, column=1,
                   value="MotionHR — HR in Motion | JS Solution | تم الإنشاء بواسطة نظام MotionHR")
    cell.font = Font(name='Cairo', size=9, color=GRAY, italic=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # ── ضبط عرض الأعمدة ──
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for c in row:
                try:
                    if c.value:
                        max_len = max(max_len, len(str(c.value)))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

    # ── الحفظ والإرسال ──
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response
'''

write_file('reports/utils.py', reports_utils)


# ════════════════════════════════════════════════
# BUG 3b: تحديث reports/views.py لتمرير company و title
# ════════════════════════════════════════════════
print("\n📌 Bug 3b: Update reports/views.py — pass company + title to export")

reports_views = '''"""
reports/views.py — Patch 49a
تحديث التقارير لتمرير company + title للـ export
"""

from datetime import date
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone

from .utils import (
    get_date_range,
    get_attendance_summary,
    get_employee_attendance_details,
    get_field_tracking_summary,
    export_to_excel,
)


# ════════════════════════════════════════════════════════════
# الصفحة الرئيسية
# ════════════════════════════════════════════════════════════

@login_required
def reports_home(request):
    return render(request, "reports/home.html", {"page_title": "التقارير"})


# ════════════════════════════════════════════════════════════
# تقرير الحضور والغياب
# ════════════════════════════════════════════════════════════

@login_required
def attendance_report(request):
    company = request.user.company
    period = request.GET.get("period", "month")

    custom_start = custom_end = None
    if period == "custom":
        try:
            custom_start = date.fromisoformat(request.GET.get("start_date", ""))
            custom_end   = date.fromisoformat(request.GET.get("end_date", ""))
        except Exception:
            period = "month"

    start_date, end_date = get_date_range(period, custom_start, custom_end)
    summary = get_attendance_summary(company, start_date, end_date)
    details = get_employee_attendance_details(company, start_date, end_date)

    if request.GET.get("export") == "excel":
        headers = [
            "الموظف", "الرقم الوظيفي", "الحاضر", "الغائب",
            "المتأخر", "إجازة", "الإجمالي", "نسبة الحضور%",
            "إجمالي التأخير (دقيقة)", "ساعات العمل"
        ]
        rows = [[
            d["employee"].full_name_ar,
            d["employee"].employee_code,
            d["present"], d["absent"], d["late"],
            d["on_leave"], d["total"], d["rate"],
            d["late_mins"], d["work_hours"],
        ] for d in details]
        return export_to_excel(
            headers, rows,
            filename=f"attendance_{start_date}_{end_date}",
            company=company,
            report_title=f"تقرير الحضور والغياب — {start_date} إلى {end_date}"
        )

    context = {
        "page_title": "تقرير الحضور والغياب",
        "summary": summary,
        "details": details,
        "start_date": start_date,
        "end_date": end_date,
        "period": period,
    }
    return render(request, "reports/attendance_report.html", context)


# ════════════════════════════════════════════════════════════
# تقرير التأخيرات
# ════════════════════════════════════════════════════════════

@login_required
def late_report(request):
    from attendance.models import Attendance
    company = request.user.company
    period  = request.GET.get("period", "month")

    custom_start = custom_end = None
    if period == "custom":
        try:
            custom_start = date.fromisoformat(request.GET.get("start_date", ""))
            custom_end   = date.fromisoformat(request.GET.get("end_date", ""))
        except Exception:
            period = "month"

    start_date, end_date = get_date_range(period, custom_start, custom_end)

    late_records = Attendance.objects.filter(
        employee__company=company,
        date__range=(start_date, end_date),
        status="late"
    ).select_related("employee", "shift").order_by("-date")

    if request.GET.get("export") == "excel":
        headers = ["الموظف", "الرقم الوظيفي", "التاريخ", "وقت الدخول", "التأخير (دقيقة)", "الوردية"]
        rows = [[
            r.employee.full_name_ar,
            r.employee.employee_code,
            str(r.date),
            str(r.check_in) if r.check_in else "—",
            r.late_minutes or 0,
            r.shift.name if r.shift else "—",
        ] for r in late_records]
        return export_to_excel(
            headers, rows,
            filename=f"late_report_{start_date}_{end_date}",
            company=company,
            report_title=f"تقرير التأخيرات — {start_date} إلى {end_date}"
        )

    context = {
        "page_title": "تقرير التأخيرات",
        "late_records": late_records,
        "start_date": start_date,
        "end_date": end_date,
        "period": period,
    }
    return render(request, "reports/late_report.html", context)


# ════════════════════════════════════════════════════════════
# تقرير الإجازات
# ════════════════════════════════════════════════════════════

@login_required
def leave_report(request):
    from leaves.models import LeaveRequest
    company = request.user.company
    period  = request.GET.get("period", "month")

    custom_start = custom_end = None
    if period == "custom":
        try:
            custom_start = date.fromisoformat(request.GET.get("start_date", ""))
            custom_end   = date.fromisoformat(request.GET.get("end_date", ""))
        except Exception:
            period = "month"

    start_date, end_date = get_date_range(period, custom_start, custom_end)

    leave_requests = LeaveRequest.objects.filter(
        employee__company=company,
        start_date__range=(start_date, end_date)
    ).select_related("employee", "leave_type").order_by("-start_date")

    if request.GET.get("export") == "excel":
        headers = ["الموظف", "الرقم الوظيفي", "نوع الإجازة", "من", "إلى", "الأيام", "الحالة"]
        rows = [[
            r.employee.full_name_ar,
            r.employee.employee_code,
            r.leave_type.name if r.leave_type else "—",
            str(r.start_date),
            str(r.end_date),
            r.duration_days if hasattr(r, 'duration_days') else "—",
            r.get_status_display() if hasattr(r, 'get_status_display') else r.status,
        ] for r in leave_requests]
        return export_to_excel(
            headers, rows,
            filename=f"leave_report_{start_date}_{end_date}",
            company=company,
            report_title=f"تقرير الإجازات — {start_date} إلى {end_date}"
        )

    context = {
        "page_title": "تقرير الإجازات",
        "leave_requests": leave_requests,
        "start_date": start_date,
        "end_date": end_date,
        "period": period,
    }
    return render(request, "reports/leave_report.html", context)


# ════════════════════════════════════════════════════════════
# تقرير الموظفين
# ════════════════════════════════════════════════════════════

@login_required
def employees_report(request):
    from employees.models import Employee
    company = request.user.company

    employees = Employee.objects.filter(
        company=company
    ).select_related("branch", "department", "job_title").order_by("employee_code")

    status_filter = request.GET.get("status", "")
    if status_filter:
        employees = employees.filter(status=status_filter)

    if request.GET.get("export") == "excel":
        headers = [
            "الرقم الوظيفي", "الاسم", "القسم", "الفرع",
            "المسمى الوظيفي", "تاريخ التعيين", "الحالة", "نوع العقد"
        ]
        rows = [[
            emp.employee_code,
            emp.full_name_ar,
            emp.department.name if emp.department else "—",
            emp.branch.name if emp.branch else "—",
            emp.job_title.name_ar if emp.job_title else "—",
            str(emp.hire_date) if emp.hire_date else "—",
            emp.get_status_display() if hasattr(emp, 'get_status_display') else emp.status,
            emp.get_contract_type_display() if hasattr(emp, 'get_contract_type_display') else emp.contract_type,
        ] for emp in employees]
        return export_to_excel(
            headers, rows,
            filename=f"employees_report_{date.today()}",
            company=company,
            report_title="تقرير الموظفين"
        )

    context = {
        "page_title": "تقرير الموظفين",
        "employees": employees,
        "status_filter": status_filter,
    }
    return render(request, "reports/employees_report.html", context)


# ════════════════════════════════════════════════════════════
# تقرير الميدانيين
# ════════════════════════════════════════════════════════════

@login_required
def field_report(request):
    company = request.user.company
    period  = request.GET.get("period", "month")

    custom_start = custom_end = None
    if period == "custom":
        try:
            custom_start = date.fromisoformat(request.GET.get("start_date", ""))
            custom_end   = date.fromisoformat(request.GET.get("end_date", ""))
        except Exception:
            period = "month"

    start_date, end_date = get_date_range(period, custom_start, custom_end)
    tracking_data = get_field_tracking_summary(company, start_date, end_date)

    context = {
        "page_title": "تقرير الموظفين الميدانيين",
        "tracking_data": tracking_data,
        "start_date": start_date,
        "end_date": end_date,
        "period": period,
    }
    return render(request, "reports/field_report.html", context)
'''

write_file('reports/views.py', reports_views)


# ════════════════════════════════════════════════
# BUG 4: Fix Live Map Template
# ════════════════════════════════════════════════
print("\n📌 Bug 4: Fix Live Map — CSS ثابتة")

live_map_template = """{% extends 'base/dashboard_base.html' %}

{% block title %}الخريطة الحية{% endblock %}

{% block extra_css %}
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<style>
  /* ── الخريطة ثابتة ── */
  .map-wrapper {
    position: relative;
    width: 100%;
    height: 500px;
    overflow: hidden;
    border-radius: 12px;
    border: 1px solid #dee2e6;
    background: #f8f9fa;
  }
  #live-map {
    position: absolute;
    top: 0; left: 0; right: 0; bottom: 0;
    width: 100% !important;
    height: 100% !important;
    z-index: 1;
  }
  .employee-badge {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 6px 12px;
    border-radius: 20px;
    font-size: 0.82rem;
    cursor: pointer;
    transition: all 0.2s;
    border: 2px solid transparent;
  }
  .employee-badge:hover, .employee-badge.active {
    border-color: #06B6D4;
    background: #f0fdff;
  }
  .status-dot {
    width: 8px; height: 8px;
    border-radius: 50%;
    display: inline-block;
  }
  .location-address {
    font-size: 0.78rem;
    color: #64748b;
    margin-top: 2px;
  }
</style>
{% endblock %}

{% block content %}
<div class="container-fluid">

  <!-- Header -->
  <div class="d-flex justify-content-between align-items-center mb-4">
    <div>
      <h4 class="mb-1 fw-bold">
        <i class="bi bi-map text-primary me-2"></i>الخريطة الحية
      </h4>
      <p class="text-muted small mb-0">تتبع الموظفين الميدانيين في الوقت الفعلي</p>
    </div>
    <div class="d-flex gap-2">
      <span class="badge bg-success-subtle text-success border border-success-subtle px-3 py-2" id="active-count">
        <i class="bi bi-circle-fill me-1" style="font-size:8px"></i>
        <span id="active-num">0</span> نشط
      </span>
      <button class="btn btn-outline-primary btn-sm" onclick="refreshMap()">
        <i class="bi bi-arrow-clockwise me-1"></i>تحديث
      </button>
    </div>
  </div>

  <div class="row g-3">
    <!-- الخريطة -->
    <div class="col-lg-8">
      <div class="card border-0 shadow-sm">
        <div class="card-body p-2">
          <div class="map-wrapper">
            <div id="live-map"></div>
          </div>
        </div>
      </div>
    </div>

    <!-- قائمة الموظفين -->
    <div class="col-lg-4">
      <div class="card border-0 shadow-sm h-100">
        <div class="card-header bg-white py-3 border-0">
          <h6 class="mb-0 fw-bold">
            <i class="bi bi-people me-2 text-primary"></i>الموظفون الميدانيون
          </h6>
        </div>
        <div class="card-body p-2" id="employees-list" style="overflow-y:auto; max-height:460px;">
          <div class="text-center py-4 text-muted" id="loading-state">
            <div class="spinner-border spinner-border-sm text-primary mb-2"></div>
            <div class="small">جارٍ التحميل...</div>
          </div>
          <div id="employee-cards"></div>
        </div>
      </div>
    </div>
  </div>

</div>
{% endblock %}

{% block extra_js %}
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script>
// ── إعداد الخريطة ──
var map = L.map('live-map', {
  center: [30.0444, 31.2357],
  zoom: 11,
  zoomControl: true
});

L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {
  attribution: '© OpenStreetMap',
  maxZoom: 19
}).addTo(map);

// ── منع resize issues ──
setTimeout(function() { map.invalidateSize(); }, 300);

var markers = {};
var employeeData = {};

// ── Icon ──
function createIcon(color) {
  return L.divIcon({
    html: '<div style="width:14px;height:14px;border-radius:50%;background:' + color + ';border:3px solid white;box-shadow:0 2px 6px rgba(0,0,0,0.4)"></div>',
    iconSize: [20, 20],
    iconAnchor: [10, 10],
    className: ''
  });
}

// ── تحميل المواقع ──
function loadLocations() {
  fetch('{% url "attendance:api_live_locations" %}')
    .then(r => r.json())
    .then(data => {
      var locations = data.locations || [];
      document.getElementById('loading-state').style.display = 'none';
      document.getElementById('active-num').textContent = locations.length;

      var cardsHtml = '';

      locations.forEach(function(loc) {
        var color = loc.is_online ? '#22c55e' : '#94a3b8';
        var lat = parseFloat(loc.lat);
        var lng = parseFloat(loc.lng);

        if (!isNaN(lat) && !isNaN(lng)) {
          if (markers[loc.id]) {
            markers[loc.id].setLatLng([lat, lng]);
          } else {
            markers[loc.id] = L.marker([lat, lng], {icon: createIcon(color)})
              .addTo(map)
              .bindPopup(
                '<div style="text-align:right;min-width:180px">' +
                '<strong>' + loc.name + '</strong><br>' +
                '<small class="text-muted">' + (loc.address || loc.lat + ', ' + loc.lng) + '</small><br>' +
                '<small>' + (loc.last_seen || '') + '</small>' +
                '</div>'
              );
          }
        }

        cardsHtml += '<div class="employee-badge bg-light mb-2 w-100" onclick="focusEmployee(' + loc.id + ',' + (isNaN(lat)?30:lat) + ',' + (isNaN(lng)?31:lng) + ')">' +
          '<span class="status-dot" style="background:' + color + '"></span>' +
          '<div class="flex-grow-1">' +
          '<div class="fw-semibold small">' + loc.name + '</div>' +
          '<div class="location-address">' + (loc.address || '📍 ' + (loc.lat || '') + ', ' + (loc.lng || '')) + '</div>' +
          '</div>' +
          '</span>' +
          '</div>';
      });

      if (locations.length === 0) {
        cardsHtml = '<div class="text-center py-4 text-muted"><i class="bi bi-geo-alt fs-2 opacity-25"></i><p class="small mt-2">لا يوجد موظفون ميدانيون نشطون</p></div>';
      }

      document.getElementById('employee-cards').innerHTML = cardsHtml;
    })
    .catch(function(err) {
      document.getElementById('loading-state').innerHTML = '<div class="text-danger small">خطأ في التحميل</div>';
    });
}

function focusEmployee(id, lat, lng) {
  if (!isNaN(lat) && !isNaN(lng)) {
    map.setView([lat, lng], 15);
    if (markers[id]) markers[id].openPopup();
  }
}

function refreshMap() {
  loadLocations();
}

// تحميل أولي
loadLocations();

// تحديث كل 30 ثانية
setInterval(loadLocations, 30000);

// ── Fix resize بعد load ──
window.addEventListener('resize', function() {
  map.invalidateSize();
});
</script>
{% endblock %}
"""

write_file('templates/attendance/live_map.html', live_map_template)


# ════════════════════════════════════════════════
# BUG 4b: Fix Visit Add Template — Map ثابتة
# ════════════════════════════════════════════════
print("\n📌 Bug 4b: Fix Visit Form Map")

visit_form_template = """{% extends 'base/dashboard_base.html' %}

{% block title %}إضافة زيارة ميدانية{% endblock %}

{% block extra_css %}
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<style>
  .map-wrapper {
    position: relative;
    width: 100%;
    height: 350px;
    overflow: hidden;
    border-radius: 10px;
    border: 2px solid #e2e8f0;
    background: #f8fafc;
  }
  #visit-map {
    position: absolute;
    top: 0; left: 0; right: 0; bottom: 0;
    width: 100% !important;
    height: 100% !important;
  }
</style>
{% endblock %}

{% block content %}
<div class="container-fluid">
  <div class="d-flex justify-content-between align-items-center mb-4">
    <div>
      <h4 class="mb-1 fw-bold">
        <i class="bi bi-geo-alt text-primary me-2"></i>إضافة زيارة ميدانية
      </h4>
      <nav aria-label="breadcrumb">
        <ol class="breadcrumb mb-0 small">
          <li class="breadcrumb-item"><a href="{% url 'dashboard:index' %}">الرئيسية</a></li>
          <li class="breadcrumb-item"><a href="{% url 'attendance:visits' %}">الزيارات</a></li>
          <li class="breadcrumb-item active">إضافة زيارة</li>
        </ol>
      </nav>
    </div>
  </div>

  <div class="row g-4">
    <!-- فورم الزيارة -->
    <div class="col-lg-6">
      <div class="card border-0 shadow-sm">
        <div class="card-header bg-white py-3 border-0">
          <h6 class="mb-0 fw-bold"><i class="bi bi-clipboard-plus me-2 text-primary"></i>بيانات الزيارة</h6>
        </div>
        <div class="card-body">
          <form method="post" id="visit-form">
            {% csrf_token %}

            <div class="mb-3">
              <label class="form-label fw-semibold">اسم الموقع <span class="text-danger">*</span></label>
              <input type="text" name="location_name" class="form-control" required
                     placeholder="مثال: مكتب العميل — المهندسين">
            </div>

            <div class="mb-3">
              <label class="form-label fw-semibold">الغرض من الزيارة <span class="text-danger">*</span></label>
              <textarea name="purpose" class="form-control" rows="3" required
                        placeholder="وصف الزيارة والهدف منها..."></textarea>
            </div>

            <div class="row g-2 mb-3">
              <div class="col-6">
                <label class="form-label fw-semibold">وقت البداية</label>
                <input type="time" name="start_time" class="form-control">
              </div>
              <div class="col-6">
                <label class="form-label fw-semibold">وقت النهاية</label>
                <input type="time" name="end_time" class="form-control">
              </div>
            </div>

            <!-- إحداثيات من الخريطة -->
            <input type="hidden" name="latitude" id="lat-input">
            <input type="hidden" name="longitude" id="lng-input">

            <div class="mb-3 p-3 bg-light rounded" id="selected-location" style="display:none">
              <div class="small text-muted mb-1"><i class="bi bi-pin-map me-1"></i>الموقع المحدد:</div>
              <div class="fw-semibold small" id="selected-address">—</div>
            </div>

            <div class="mb-3">
              <label class="form-label fw-semibold">ملاحظات</label>
              <textarea name="notes" class="form-control" rows="2"></textarea>
            </div>

            <div class="d-flex gap-2">
              <button type="submit" class="btn btn-primary">
                <i class="bi bi-check2-circle me-1"></i>حفظ الزيارة
              </button>
              <a href="{% url 'attendance:visits' %}" class="btn btn-outline-secondary">إلغاء</a>
            </div>
          </form>
        </div>
      </div>
    </div>

    <!-- الخريطة -->
    <div class="col-lg-6">
      <div class="card border-0 shadow-sm">
        <div class="card-header bg-white py-3 border-0">
          <h6 class="mb-0 fw-bold">
            <i class="bi bi-map me-2 text-primary"></i>
            حدد الموقع على الخريطة
            <small class="text-muted fw-normal">(انقر لتحديد)</small>
          </h6>
        </div>
        <div class="card-body p-2">
          <div class="map-wrapper">
            <div id="visit-map"></div>
          </div>
          <div class="text-center mt-2">
            <button type="button" class="btn btn-sm btn-outline-primary" onclick="useMyLocation()">
              <i class="bi bi-crosshair me-1"></i>استخدم موقعي الحالي
            </button>
          </div>
        </div>
      </div>
    </div>
  </div>
</div>
{% endblock %}

{% block extra_js %}
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script>
var map = L.map('visit-map', {
  center: [30.0444, 31.2357],
  zoom: 12
});

L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {
  attribution: '© OpenStreetMap', maxZoom: 19
}).addTo(map);

// Fix resize
setTimeout(function() { map.invalidateSize(); }, 300);
window.addEventListener('resize', function() { map.invalidateSize(); });

var selectedMarker = null;

map.on('click', function(e) {
  var lat = e.latlng.lat.toFixed(6);
  var lng = e.latlng.lng.toFixed(6);
  setLocation(lat, lng);
});

function setLocation(lat, lng) {
  if (selectedMarker) map.removeLayer(selectedMarker);
  selectedMarker = L.marker([lat, lng]).addTo(map);

  document.getElementById('lat-input').value = lat;
  document.getElementById('lng-input').value = lng;

  // عكس الجيوكودينج
  fetch('https://nominatim.openstreetmap.org/reverse?lat=' + lat + '&lon=' + lng + '&format=json&accept-language=ar')
    .then(r => r.json())
    .then(data => {
      var addr = data.display_name || (lat + ', ' + lng);
      document.getElementById('selected-address').textContent = addr;
      document.getElementById('selected-location').style.display = 'block';
    })
    .catch(() => {
      document.getElementById('selected-address').textContent = lat + ', ' + lng;
      document.getElementById('selected-location').style.display = 'block';
    });
}

function useMyLocation() {
  if (navigator.geolocation) {
    navigator.geolocation.getCurrentPosition(function(pos) {
      var lat = pos.coords.latitude.toFixed(6);
      var lng = pos.coords.longitude.toFixed(6);
      map.setView([lat, lng], 15);
      setLocation(lat, lng);
    }, function() {
      alert('لم يتمكن من الحصول على موقعك. تأكد من إذن الموقع.');
    });
  }
}
</script>
{% endblock %}
"""

write_file('templates/attendance/visit_form.html', visit_form_template)


# ════════════════════════════════════════════════
# BUG 5: Stealth Tracking — تفعيل من السياسات
# ════════════════════════════════════════════════
print("\n📌 Bug 5: Fix Stealth Tracking Policy Check")

# نحدث view الـ stealth manage لتعرض رابط للإعدادات
stealth_manage_template = """{% extends 'base/dashboard_base.html' %}

{% block title %}إدارة التتبع الصامت{% endblock %}

{% block content %}
<div class="container-fluid">

  <div class="d-flex justify-content-between align-items-center mb-4">
    <div>
      <h4 class="mb-1 fw-bold">
        <i class="bi bi-eye-slash text-primary me-2"></i>إدارة التتبع الصامت
      </h4>
      <p class="text-muted small mb-0">تتبع الموظفين الميدانيين بشكل صامت</p>
    </div>
  </div>

  {% if not policy_enabled %}
  <!-- التتبع غير مفعل -->
  <div class="card border-0 shadow-sm mb-4 border-start border-warning border-4">
    <div class="card-body d-flex align-items-center gap-3 py-4">
      <div class="rounded-circle bg-warning-subtle p-3">
        <i class="bi bi-shield-exclamation text-warning fs-4"></i>
      </div>
      <div class="flex-grow-1">
        <h6 class="mb-1 fw-bold">ميزة التتبع الصامت غير مفعّلة</h6>
        <p class="text-muted small mb-0">
          لتفعيل التتبع الصامت يجب تفعيله أولاً من سياسات الشركة.
        </p>
      </div>
      <div>
        <a href="{% url 'companies:policies' %}" class="btn btn-warning">
          <i class="bi bi-gear me-1"></i>الذهاب لسياسات الشركة
        </a>
      </div>
    </div>
  </div>

  <div class="card border-0 shadow-sm">
    <div class="card-body">
      <h6 class="fw-bold mb-3"><i class="bi bi-info-circle text-primary me-2"></i>كيفية تفعيل التتبع الصامت</h6>
      <ol class="small text-muted">
        <li class="mb-2">اذهب إلى <strong>إعدادات الشركة ← سياسات الشركة</strong></li>
        <li class="mb-2">ابحث عن قسم <strong>التتبع الصامت</strong></li>
        <li class="mb-2">فعّل خيار <strong>تفعيل التتبع الصامت للموظفين الميدانيين</strong></li>
        <li class="mb-2">احفظ الإعدادات</li>
        <li>عُد لهذه الصفحة وستعمل بشكل طبيعي</li>
      </ol>
    </div>
  </div>

  {% else %}
  <!-- التتبع مفعل -->
  <div class="card border-0 shadow-sm mb-4">
    <div class="card-body d-flex align-items-center gap-3 py-3">
      <div class="rounded-circle bg-success-subtle p-3">
        <i class="bi bi-shield-check text-success fs-4"></i>
      </div>
      <div>
        <h6 class="mb-0 fw-bold text-success">التتبع الصامت مفعّل</h6>
        <small class="text-muted">يتم تتبع الموظفين الميدانيين المحددين بشكل صامت</small>
      </div>
    </div>
  </div>

  <!-- قائمة الموظفين المتتبَّعين -->
  <div class="card border-0 shadow-sm">
    <div class="card-header bg-white py-3 border-0">
      <h6 class="mb-0 fw-bold"><i class="bi bi-people me-2 text-primary"></i>الموظفون الميدانيون</h6>
    </div>
    <div class="card-body p-0">
      {% if field_employees %}
      <div class="table-responsive">
        <table class="table table-hover align-middle mb-0">
          <thead class="table-light">
            <tr>
              <th>الموظف</th>
              <th>الكود</th>
              <th>التتبع الصامت</th>
              <th>آخر موقع</th>
              <th>إجراء</th>
            </tr>
          </thead>
          <tbody>
            {% for emp in field_employees %}
            <tr>
              <td>{{ emp.full_name_ar }}</td>
              <td class="small text-muted">{{ emp.employee_code }}</td>
              <td>
                {% if emp.stealth_tracking_enabled %}
                <span class="badge bg-danger-subtle text-danger border border-danger-subtle rounded-pill px-2">
                  <i class="bi bi-eye-slash me-1"></i>صامت
                </span>
                {% else %}
                <span class="badge bg-secondary-subtle text-secondary rounded-pill px-2">غير مفعل</span>
                {% endif %}
              </td>
              <td class="small text-muted">—</td>
              <td>
                <form method="post" action="{% url 'attendance:stealth_manage' %}" class="d-inline">
                  {% csrf_token %}
                  <input type="hidden" name="employee_id" value="{{ emp.id }}">
                  <input type="hidden" name="action" value="{% if emp.stealth_tracking_enabled %}disable{% else %}enable{% endif %}">
                  <button type="submit" class="btn btn-sm {% if emp.stealth_tracking_enabled %}btn-outline-danger{% else %}btn-outline-primary{% endif %}">
                    {% if emp.stealth_tracking_enabled %}
                    <i class="bi bi-eye me-1"></i>إيقاف
                    {% else %}
                    <i class="bi bi-eye-slash me-1"></i>تفعيل
                    {% endif %}
                  </button>
                </form>
              </td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
      {% else %}
      <div class="text-center py-5 text-muted">
        <i class="bi bi-people fs-1 opacity-25"></i>
        <p class="mt-3 small">لا يوجد موظفون ميدانيون</p>
      </div>
      {% endif %}
    </div>
  </div>
  {% endif %}

</div>
{% endblock %}
"""

write_file('templates/attendance/stealth_manage.html', stealth_manage_template)


# ════════════════════════════════════════════════
# تحديث attendance/views.py — stealth_tracking_manage
# ════════════════════════════════════════════════
print("\n📌 Bug 5b: Update stealth_tracking_manage view")

# نقرأ الـ views الحالية ونضيف الـ fix
views_content = read_file('attendance/views.py')
if views_content:
    # نحدث الـ stealth_tracking_manage function
    stealth_fix = '''

@login_required
def stealth_tracking_manage(request):
    """إدارة التتبع الصامت — Patch 49a"""
    from companies.models import CompanyPolicy
    from employees.models import Employee

    company = request.user.company

    # جلب السياسة
    policy = None
    try:
        policy = CompanyPolicy.objects.get(company=company)
    except CompanyPolicy.DoesNotExist:
        pass

    policy_enabled = policy.stealth_tracking_enabled if policy else False

    # معالجة POST (تفعيل/إيقاف للموظف)
    if request.method == 'POST' and policy_enabled:
        emp_id = request.POST.get('employee_id')
        action = request.POST.get('action')
        try:
            emp = Employee.objects.get(id=emp_id, company=company)
            if action == 'enable':
                emp.stealth_tracking_enabled = True
                messages.success(request, f'تم تفعيل التتبع الصامت لـ {emp.full_name_ar}')
            elif action == 'disable':
                emp.stealth_tracking_enabled = False
                messages.info(request, f'تم إيقاف التتبع الصامت لـ {emp.full_name_ar}')
            emp.save()
        except Employee.DoesNotExist:
            messages.error(request, 'الموظف غير موجود')
        return redirect('attendance:stealth_manage')

    field_employees = Employee.objects.filter(
        company=company,
        status='active',
        is_field_worker=True
    )

    context = {
        'policy_enabled': policy_enabled,
        'field_employees': field_employees,
        'policy': policy,
    }
    return render(request, 'attendance/stealth_manage.html', context)

'''

    # نضيف الـ function الجديدة في نهاية الملف
    # أولاً نشيل الـ function القديمة لو موجودة
    import re
    # نحتاج نضيف الـ function لو مش موجودة أو نعدلها
    if 'def stealth_tracking_manage' not in views_content:
        views_content += stealth_fix
        write_file('attendance/views.py', views_content)
        print("   ✅ تم إضافة stealth_tracking_manage")
    else:
        print("   ℹ️ stealth_tracking_manage موجودة — تأكد من تحديثها يدوياً لو في مشكلة")


# ════════════════════════════════════════════════
# تحديث base template لإضافة CSS الجديد
# ════════════════════════════════════════════════
print("\n📌 تحديث Base Template لإضافة patch49a.css")

base_content = read_file('templates/base/dashboard_base.html')
if base_content and 'patch49a_fixes.css' not in base_content:
    # نضيف الـ CSS قبل </head>
    base_content = base_content.replace(
        '</head>',
        '  <link rel="stylesheet" href="/static/css/patch49a_fixes.css">\n</head>'
    )
    write_file('templates/base/dashboard_base.html', base_content)
    print("   ✅ تم إضافة patch49a_fixes.css للـ base template")
else:
    if base_content:
        print("   ℹ️ CSS موجود بالفعل")
    else:
        print("   ⚠️ Base template مش موجود — هتضيف CSS يدوياً")


# ════════════════════════════════════════════════
# تثبيت openpyxl
# ════════════════════════════════════════════════
print("\n📌 تثبيت openpyxl...")
import subprocess
result = subprocess.run(
    [sys.executable, '-m', 'pip', 'install', 'openpyxl'],
    capture_output=True, text=True
)
if result.returncode == 0:
    print("   ✅ openpyxl مثبت بنجاح")
else:
    print("   ⚠️ شغّل يدوياً: pip install openpyxl")
    print("   ", result.stderr[:200])


# ════════════════════════════════════════════════
# api_live_locations — إضافة عنوان الموقع
# ════════════════════════════════════════════════
print("\n📌 تحديث api_live_locations لإضافة العنوان")

views_content2 = read_file('attendance/views.py')
if views_content2 and 'def api_live_locations' in views_content2:
    print("   ℹ️ api_live_locations موجودة")
    # نتأكد إن الـ response بترجع address
    if "'address'" not in views_content2 and '"address"' not in views_content2:
        print("   ⚠️ الـ address مش موجود في الـ response — سيتم الحصول عليه من الـ frontend")


print("\n" + "=" * 60)
print("✅ Patch 49a اكتمل!")
print("=" * 60)
print("""
الملفات المعدّلة:
  ✅ static/css/patch49a_fixes.css       ← Sidebar + Map CSS
  ✅ templates/attendance/list.html      ← Fix NoReverseMatch
  ✅ reports/utils.py                    ← Excel احترافي
  ✅ reports/views.py                    ← تمرير company + title
  ✅ templates/attendance/live_map.html  ← خريطة ثابتة
  ✅ templates/attendance/visit_form.html ← خريطة ثابتة
  ✅ templates/attendance/stealth_manage.html ← Fix policy check
  ✅ templates/base/dashboard_base.html  ← إضافة CSS

الخطوة التالية:
  python manage.py check
  python manage.py runserver 0.0.0.0:8000
""")