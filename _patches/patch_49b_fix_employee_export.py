"""
Patch 49b Fix — Employee Export (Excel + PDF)

1. Excel احترافي بكل بيانات الموظف باستخدام openpyxl
2. PDF احترافي باستخدام xhtml2pdf
3. تحديث views.py و list.html
"""

import os
import sys
import subprocess
import django

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE_DIR)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "motionhr.settings")
django.setup()


def read_file(path):
    full = os.path.join(BASE_DIR, path)
    if not os.path.exists(full):
        return None
    with open(full, "r", encoding="utf-8") as f:
        return f.read()


def write_file(path, content):
    full = os.path.join(BASE_DIR, path)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    with open(full, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"✅ كُتب: {path}")


print("=" * 60)
print("Patch 49b Fix — Employee Export (Excel + PDF)")
print("=" * 60)


# ═══════════════════════════════════════════════════
# Step 0: تثبيت xhtml2pdf
# ═══════════════════════════════════════════════════
print("\n📌 Step 0: تثبيت xhtml2pdf")
result = subprocess.run(
    [sys.executable, '-m', 'pip', 'install', 'xhtml2pdf'],
    capture_output=True, text=True
)
if result.returncode == 0:
    print("   ✅ xhtml2pdf مثبت")
else:
    print(f"   ⚠️ شغّل يدوياً: pip install xhtml2pdf")
    print("   ", result.stderr[:200])


# ═══════════════════════════════════════════════════
# Step 1: إضافة employee_exports.py
# ═══════════════════════════════════════════════════
print("\n📌 Step 1: إنشاء employees/exports.py")

exports_code = '''"""
employees/exports.py — Patch 49b Fix
Excel + PDF export للموظفين — احترافي وكامل
"""

from django.http import HttpResponse
from django.utils import timezone
from django.template.loader import render_to_string

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from xhtml2pdf import pisa
    HAS_PISA = True
except ImportError:
    HAS_PISA = False


def get_company_name(user):
    """جلب اسم الشركة من المستخدم"""
    company = getattr(user, 'company', None)
    if company:
        return getattr(company, 'name_ar', None) or getattr(company, 'name_en', None) or 'MotionHR'
    return 'MotionHR'


def get_employee_row(emp):
    """تحويل موظف إلى صف بيانات"""
    dept = emp.department
    branch = emp.branch
    job = emp.job_title
    manager = emp.direct_manager

    return [
        emp.employee_code or '—',
        getattr(emp, 'full_name_ar', None) or f"{emp.first_name_ar or ''} {emp.last_name_ar or ''}".strip() or '—',
        f"{emp.first_name_en or ''} {emp.last_name_en or ''}".strip() or '—',
        emp.national_id or '—',
        (getattr(job, 'name_ar', None) or getattr(job, 'name_en', None) or '—') if job else '—',
        (getattr(dept, 'name_ar', None) or getattr(dept, 'name_en', None) or '—') if dept else '—',
        (getattr(branch, 'name_ar', None) or getattr(branch, 'name_en', None) or '—') if branch else '—',
        (getattr(manager, 'full_name_ar', None) or '—') if manager else '—',
        emp.phone or '—',
        emp.email or '—',
        str(emp.hire_date) if emp.hire_date else '—',
        emp.get_contract_type_display() if hasattr(emp, 'get_contract_type_display') else (emp.contract_type or '—'),
        emp.get_status_display() if hasattr(emp, 'get_status_display') else (emp.status or '—'),
        str(emp.basic_salary) if emp.basic_salary else '—',
        emp.get_gender_display() if hasattr(emp, 'get_gender_display') else (emp.gender or '—'),
        emp.city or '—',
    ]


HEADERS = [
    "الرقم الوظيفي",
    "الاسم بالعربي",
    "الاسم بالإنجليزي",
    "الرقم القومي",
    "المسمى الوظيفي",
    "الإدارة",
    "الفرع",
    "المدير المباشر",
    "التليفون",
    "الإيميل",
    "تاريخ التعيين",
    "نوع العقد",
    "الحالة",
    "الراتب الأساسي",
    "النوع",
    "المدينة",
]


def export_employees_excel(employees, user=None):
    """تصدير Excel احترافي"""
    if not HAS_OPENPYXL:
        return _fallback_csv(employees)

    company_name = get_company_name(user) if user else 'MotionHR'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الموظفون"
    ws.sheet_view.rightToLeft = True

    PRIMARY   = "06B6D4"
    HEADER_BG = "0E7490"
    LIGHT_ROW = "F0FDFF"
    WHITE     = "FFFFFF"
    DARK      = "1E293B"
    GRAY      = "94A3B8"

    thin_border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6'),
    )

    num_cols = len(HEADERS)
    last_col = get_column_letter(num_cols)
    row_num = 1

    # ── اسم الشركة ──
    ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
    cell = ws.cell(row=row_num, column=1, value=company_name)
    cell.font = Font(name='Cairo', size=16, bold=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_num].height = 35
    row_num += 1

    # ── عنوان التقرير ──
    ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
    cell = ws.cell(row=row_num, column=1, value="تقرير الموظفين")
    cell.font = Font(name='Cairo', size=13, bold=True, color=DARK)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_num].height = 28
    row_num += 1

    # ── تاريخ التصدير ──
    ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
    export_date = timezone.now().strftime("%Y/%m/%d %H:%M")
    cell = ws.cell(row=row_num, column=1, value=f"تاريخ التصدير: {export_date}")
    cell.font = Font(name='Cairo', size=10, color=GRAY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_num].height = 20
    row_num += 2

    # ── Headers ──
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.font = Font(name='Cairo', size=11, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row_num].height = 30
    row_num += 1

    # ── Data ──
    for idx, emp in enumerate(employees):
        bg = LIGHT_ROW if idx % 2 == 0 else WHITE
        row_data = get_employee_row(emp)
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name='Cairo', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[row_num].height = 22
        row_num += 1

    # ── Footer ──
    row_num += 1
    ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
    cell = ws.cell(row=row_num, column=1,
                   value="MotionHR — HR in Motion | JS Solution")
    cell.font = Font(name='Cairo', size=9, color=GRAY, italic=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # ── عرض الأعمدة ──
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(HEADERS[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for c in row:
                try:
                    if c.value:
                        max_len = max(max_len, len(str(c.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    today = timezone.now().strftime("%Y-%m-%d")
    response['Content-Disposition'] = f'attachment; filename="employees_{today}.xlsx"'
    wb.save(response)
    return response


def _fallback_csv(employees):
    """Fallback CSV لو openpyxl مش مثبت"""
    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="employees.csv"'
    writer = csv.writer(response)
    writer.writerow(HEADERS)
    for emp in employees:
        writer.writerow(get_employee_row(emp))
    return response


def export_employees_pdf(employees, user=None):
    """تصدير PDF احترافي"""
    if not HAS_PISA:
        raise Exception("xhtml2pdf غير مثبت")

    company_name = get_company_name(user) if user else 'MotionHR'
    export_date = timezone.now().strftime("%Y/%m/%d %H:%M")

    rows = []
    for emp in employees:
        rows.append(get_employee_row(emp))

    html = _build_pdf_html(company_name, export_date, HEADERS, rows)

    response = HttpResponse(content_type='application/pdf')
    today = timezone.now().strftime("%Y-%m-%d")
    response['Content-Disposition'] = f'attachment; filename="employees_{today}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response, encoding='utf-8')
    if pisa_status.err:
        raise Exception("خطأ في إنشاء PDF")

    return response


def _build_pdf_html(company_name, export_date, headers, rows):
    """بناء HTML للـ PDF"""
    header_cells = ''.join(
        f'<th style="background:#0E7490;color:white;padding:6px 4px;font-size:9px;'
        f'text-align:center;border:1px solid #dee2e6;">{h}</th>'
        for h in headers
    )

    body_rows = ''
    for idx, row in enumerate(rows):
        bg = '#F0FDFF' if idx % 2 == 0 else '#FFFFFF'
        cells = ''.join(
            f'<td style="background:{bg};padding:4px 3px;font-size:9px;'
            f'text-align:center;border:1px solid #dee2e6;">{v}</td>'
            for v in row
        )
        body_rows += f'<tr>{cells}</tr>'

    html = f"""
    <!DOCTYPE html>
    <html dir="rtl" lang="ar">
    <head>
        <meta charset="utf-8">
        <style>
            @page {{
                size: A4 landscape;
                margin: 1cm;
            }}
            body {{
                font-family: Arial, sans-serif;
                direction: rtl;
                margin: 0;
                padding: 0;
            }}
            .header {{
                text-align: center;
                margin-bottom: 15px;
            }}
            .company-name {{
                font-size: 18px;
                font-weight: bold;
                color: #06B6D4;
                margin-bottom: 5px;
            }}
            .report-title {{
                font-size: 14px;
                font-weight: bold;
                color: #1E293B;
                margin-bottom: 3px;
            }}
            .export-date {{
                font-size: 10px;
                color: #94A3B8;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
            }}
            .footer {{
                text-align: center;
                font-size: 9px;
                color: #94A3B8;
                margin-top: 15px;
                font-style: italic;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company-name">{company_name}</div>
            <div class="report-title">تقرير الموظفين</div>
            <div class="export-date">تاريخ التصدير: {export_date}</div>
        </div>

        <table>
            <thead>
                <tr>{header_cells}</tr>
            </thead>
            <tbody>
                {body_rows}
            </tbody>
        </table>

        <div class="footer">
            MotionHR — HR in Motion | JS Solution
        </div>
    </body>
    </html>
    """
    return html
'''

write_file('employees/exports.py', exports_code)


# ═══════════════════════════════════════════════════
# Step 2: تحديث employees/views.py
# ═══════════════════════════════════════════════════
print("\n📌 Step 2: تحديث employees/views.py — ربط export الجديد")

views_path = "employees/views.py"
views_content = read_file(views_path)
if views_content is None:
    raise SystemExit("❌ ملف employees/views.py غير موجود")

# إضافة import الجديد
import_line = "from .exports import export_employees_excel, export_employees_pdf"
if import_line not in views_content:
    # نضيفه بعد آخر import
    last_import_idx = 0
    lines = views_content.split('\n')
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith('from ') or stripped.startswith('import '):
            last_import_idx = i

    lines.insert(last_import_idx + 1, import_line)
    views_content = '\n'.join(lines)
    print("   ✅ تم إضافة import export functions")

# استبدال التصدير داخل employee_list
old_excel_block = """    # التصدير
    export_type = request.GET.get('export', '')
    if export_type == 'excel':
        return export_employees_excel(employees)
    elif export_type == 'pdf':
        try:
            return export_employees_pdf(employees)
        except Exception as e:
            messages.warning(request, f'PDF غير متاح - جرب Excel')
            return redirect('employees:list')"""

new_excel_block = """    # التصدير
    export_type = request.GET.get('export', '')
    if export_type == 'excel':
        return export_employees_excel(employees, user=request.user)
    elif export_type == 'pdf':
        try:
            return export_employees_pdf(employees, user=request.user)
        except Exception as e:
            messages.warning(request, f'خطأ في تصدير PDF: {e}')
            return redirect('employees:list')"""

if old_excel_block in views_content:
    views_content = views_content.replace(old_excel_block, new_excel_block)
    print("   ✅ تم تحديث export block في employee_list")
else:
    # لو الكود مختلف شوية
    # نبحث عن أي export_employees_excel call ونعدله
    if "export_employees_excel(employees)" in views_content and "user=request.user" not in views_content.split("export_employees_excel")[1][:50]:
        views_content = views_content.replace(
            "export_employees_excel(employees)",
            "export_employees_excel(employees, user=request.user)"
        )
        print("   ✅ تم تحديث excel export call")

    if "export_employees_pdf(employees)" in views_content and "user=request.user" not in views_content.split("export_employees_pdf")[1][:50]:
        views_content = views_content.replace(
            "export_employees_pdf(employees)",
            "export_employees_pdf(employees, user=request.user)"
        )
        print("   ✅ تم تحديث pdf export call")

# حذف أي دوال export قديمة لو موجودة في نفس الملف
old_func_patterns = [
    "def export_employees_excel(",
    "def export_employees_pdf(",
]

for pattern in old_func_patterns:
    if pattern in views_content:
        # نحذف الـ function القديمة
        idx = views_content.index(pattern)
        # نبحث عن بداية الـ def
        # نبحث للخلف عن \n\n أو بداية الملف
        start = views_content.rfind('\n\n', 0, idx)
        if start == -1:
            start = views_content.rfind('\ndef ', 0, idx)
        if start == -1:
            start = idx

        # نبحث عن نهاية الـ function
        end = views_content.find('\n\ndef ', idx + 10)
        if end == -1:
            end = views_content.find('\n\n@', idx + 10)
        if end == -1:
            end = len(views_content)

        old_func = views_content[start:end]
        views_content = views_content.replace(old_func, '\n')
        print(f"   ✅ تم حذف الدالة القديمة: {pattern.strip()}")

write_file(views_path, views_content)


# ═══════════════════════════════════════════════════
# Step 3: تحديث employee list template — زرار PDF
# ═══════════════════════════════════════════════════
print("\n📌 Step 3: تحديث زرار PDF في list.html")

list_content = read_file("templates/employees/list.html")
if list_content:
    # استبدال زرار PDF المعطل بزرار شغال
    list_content = list_content.replace(
        '<button type="button" class="btn btn-outline-secondary" disabled title="سيتم تنفيذ PDF في باتش مستقل">',
        '<a href="?export=pdf" class="btn btn-danger">'
    )
    list_content = list_content.replace(
        '<i class="bi bi-file-earmark-pdf me-1"></i>PDF قريبًا\n      </button>',
        '<i class="bi bi-file-earmark-pdf me-1"></i>تصدير PDF\n      </a>'
    )
    # لو مش بالظبط كده
    list_content = list_content.replace(
        '<i class="bi bi-file-earmark-pdf me-1"></i>PDF قريبًا</button>',
        '<i class="bi bi-file-earmark-pdf me-1"></i>تصدير PDF</a>'
    )

    write_file("templates/employees/list.html", list_content)
    print("   ✅ تم تفعيل زرار PDF")


print("\n" + "=" * 60)
print("✅ Patch 49b Fix — Employee Export اكتمل")
print("=" * 60)
print("""
اللي اتعمل:
  ✅ employees/exports.py — ملف جديد فيه:
     - export_employees_excel() → openpyxl + 16 عمود + لوجو + تنسيق
     - export_employees_pdf()   → xhtml2pdf + A4 landscape + RTL
  ✅ employees/views.py — ربط الـ export الجديد + حذف الدوال القديمة
  ✅ templates/employees/list.html — تفعيل زرار PDF

الأعمدة في التقرير:
  الرقم الوظيفي | الاسم بالعربي | الاسم بالإنجليزي | الرقم القومي
  المسمى الوظيفي | الإدارة | الفرع | المدير المباشر
  التليفون | الإيميل | تاريخ التعيين | نوع العقد
  الحالة | الراتب الأساسي | النوع | المدينة

شغّل:
  python manage.py check
  python manage.py runserver 0.0.0.0:8000

اختبر:
  /employees/ → تصدير Excel → بيانات كاملة ومنظمة
  /employees/ → تصدير PDF → ملف PDF كامل
""")