"""
Patch 49a Fix3 — إصلاح بناءً على الـ field names الحقيقية
1. Fix dashboard URL (مفيش namespace)
2. Fix check_in_time / check_out_time
3. Fix company.name_ar
4. Fix department.name_ar
5. Fix stealth_tracking — enable من الـ view مباشرة
"""

import os
import sys
import django

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'motionhr.settings')
django.setup()

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def write_file(path, content):
    full = os.path.join(BASE, path)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    with open(full, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"✅ كُتب: {path}")

def read_file(path):
    full = os.path.join(BASE, path)
    if not os.path.exists(full):
        return None
    with open(full, 'r', encoding='utf-8') as f:
        return f.read()

print("=" * 60)
print("Patch 49a Fix3 — الإصلاح الكامل")
print("=" * 60)


# ════════════════════════════════════════════════
# FIX 1: Attendance List Template — dashboard URL
# ════════════════════════════════════════════════
print("\n📌 Fix 1: Attendance List — Fix dashboard URL")

attendance_list_template = """{% extends 'base/dashboard_base.html' %}
{% load custom_filters %}

{% block title %}سجلات الحضور{% endblock %}

{% block content %}
<div class="container-fluid">

  <!-- Header -->
  <div class="d-flex justify-content-between align-items-center mb-4">
    <div>
      <h4 class="mb-1 fw-bold">
        <i class="bi bi-calendar-check text-primary me-2"></i>سجلات الحضور
      </h4>
      <nav aria-label="breadcrumb">
        <ol class="breadcrumb mb-0 small">
          <li class="breadcrumb-item"><a href="{% url 'dashboard' %}">الرئيسية</a></li>
          <li class="breadcrumb-item active">سجلات الحضور</li>
        </ol>
      </nav>
    </div>
    {% if request.user.role == 'company_admin' or request.user.role == 'hr_manager' %}
    <a href="{% url 'attendance:check_in' %}" class="btn btn-primary">
      <i class="bi bi-plus-circle me-1"></i>تسجيل حضور
    </a>
    {% endif %}
  </div>

  <!-- إحصائيات اليوم -->
  <div class="row g-3 mb-4">
    <div class="col-6 col-md-3">
      <div class="card border-0 shadow-sm text-center py-3">
        <div class="fs-2 fw-bold text-primary">{{ today_stats.total }}</div>
        <div class="text-muted small">إجمالي اليوم</div>
      </div>
    </div>
    <div class="col-6 col-md-3">
      <div class="card border-0 shadow-sm text-center py-3">
        <div class="fs-2 fw-bold text-success">{{ today_stats.present }}</div>
        <div class="text-muted small">حاضر</div>
      </div>
    </div>
    <div class="col-6 col-md-3">
      <div class="card border-0 shadow-sm text-center py-3">
        <div class="fs-2 fw-bold text-warning">{{ today_stats.late }}</div>
        <div class="text-muted small">متأخر</div>
      </div>
    </div>
    <div class="col-6 col-md-3">
      <div class="card border-0 shadow-sm text-center py-3">
        <div class="fs-2 fw-bold text-danger">{{ today_stats.absent }}</div>
        <div class="text-muted small">غائب</div>
      </div>
    </div>
  </div>

  <!-- فلترة -->
  <div class="card border-0 shadow-sm mb-4">
    <div class="card-body">
      <form method="get" class="row g-2 align-items-end">
        <div class="col-md-3">
          <label class="form-label small text-muted">من تاريخ</label>
          <input type="date" name="date_from" value="{{ date_from }}" class="form-control form-control-sm">
        </div>
        <div class="col-md-3">
          <label class="form-label small text-muted">إلى تاريخ</label>
          <input type="date" name="date_to" value="{{ date_to }}" class="form-control form-control-sm">
        </div>
        <div class="col-md-3">
          <label class="form-label small text-muted">الموظف</label>
          <select name="employee" class="form-select form-select-sm">
            <option value="">-- الكل --</option>
            {% for emp in employees %}
            <option value="{{ emp.id }}" {% if employee_id == emp.id|stringformat:'s' %}selected{% endif %}>
              {{ emp.full_name_ar }}
            </option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-2">
          <label class="form-label small text-muted">الحالة</label>
          <select name="status" class="form-select form-select-sm">
            <option value="">-- الكل --</option>
            {% for val, label in status_choices %}
            <option value="{{ val }}" {% if status == val %}selected{% endif %}>{{ label }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-1">
          <button type="submit" class="btn btn-primary btn-sm w-100">
            <i class="bi bi-search"></i>
          </button>
        </div>
      </form>
    </div>
  </div>

  <!-- الجدول -->
  <div class="card border-0 shadow-sm">
    <div class="card-header bg-white d-flex justify-content-between align-items-center py-3">
      <span class="fw-semibold">
        <i class="bi bi-list-ul me-2 text-primary"></i>
        السجلات ({{ total_count }})
      </span>
    </div>
    <div class="card-body p-0">
      {% if attendances %}
      <div class="table-responsive">
        <table class="table table-hover align-middle mb-0">
          <thead class="table-light">
            <tr>
              <th>الموظف</th>
              <th>التاريخ</th>
              <th>وقت الدخول</th>
              <th>وقت الخروج</th>
              <th>الحالة</th>
              <th>التأخير</th>
              <th>ملاحظات</th>
              {% if request.user.role == 'company_admin' or request.user.role == 'hr_manager' %}
              <th>إجراء</th>
              {% endif %}
            </tr>
          </thead>
          <tbody>
            {% for att in attendances %}
            <tr>
              <td>
                <div class="fw-semibold small">{{ att.employee.full_name_ar }}</div>
                <div class="text-muted" style="font-size:0.75rem">{{ att.employee.employee_code }}</div>
              </td>
              <td class="small">{{ att.date|date:"Y/m/d" }}</td>
              <td class="small">
                {% if att.check_in_time %}
                  <span class="text-success fw-semibold">{{ att.check_in_time|time:"H:i" }}</span>
                {% else %}
                  <span class="text-muted">—</span>
                {% endif %}
              </td>
              <td class="small">
                {% if att.check_out_time %}
                  <span class="text-primary fw-semibold">{{ att.check_out_time|time:"H:i" }}</span>
                {% else %}
                  <span class="text-muted">—</span>
                {% endif %}
              </td>
              <td>
                {% if att.status == 'present' %}
                  <span class="badge bg-success-subtle text-success border border-success-subtle rounded-pill px-2">حاضر</span>
                {% elif att.status == 'late' %}
                  <span class="badge bg-warning-subtle text-warning border border-warning-subtle rounded-pill px-2">متأخر</span>
                {% elif att.status == 'absent' %}
                  <span class="badge bg-danger-subtle text-danger border border-danger-subtle rounded-pill px-2">غائب</span>
                {% elif att.status == 'on_leave' %}
                  <span class="badge bg-info-subtle text-info border border-info-subtle rounded-pill px-2">إجازة</span>
                {% else %}
                  <span class="badge bg-secondary-subtle text-secondary rounded-pill px-2">{{ att.status }}</span>
                {% endif %}
              </td>
              <td class="small">
                {% if att.late_minutes and att.late_minutes > 0 %}
                  <span class="text-danger">{{ att.late_minutes }} د</span>
                {% else %}
                  <span class="text-muted">—</span>
                {% endif %}
              </td>
              <td class="small text-muted">
                {{ att.admin_notes|default:"—"|truncatechars:30 }}
              </td>
              {% if request.user.role == 'company_admin' or request.user.role == 'hr_manager' %}
              <td>
                {% if att.id %}
                <a href="{% url 'attendance:override' att.id %}"
                   class="btn btn-sm btn-outline-secondary"
                   title="تعديل">
                  <i class="bi bi-pencil"></i>
                </a>
                {% endif %}
              </td>
              {% endif %}
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>

      <!-- Pagination -->
      {% if page_obj.has_other_pages %}
      <div class="d-flex justify-content-center py-3">
        <nav>
          <ul class="pagination pagination-sm mb-0">
            {% if page_obj.has_previous %}
            <li class="page-item">
              <a class="page-link" href="?page={{ page_obj.previous_page_number }}&date_from={{ date_from }}&date_to={{ date_to }}&employee={{ employee_id }}&status={{ status }}">
                <i class="bi bi-chevron-right"></i>
              </a>
            </li>
            {% endif %}
            <li class="page-item active">
              <span class="page-link">{{ page_obj.number }} / {{ page_obj.paginator.num_pages }}</span>
            </li>
            {% if page_obj.has_next %}
            <li class="page-item">
              <a class="page-link" href="?page={{ page_obj.next_page_number }}&date_from={{ date_from }}&date_to={{ date_to }}&employee={{ employee_id }}&status={{ status }}">
                <i class="bi bi-chevron-left"></i>
              </a>
            </li>
            {% endif %}
          </ul>
        </nav>
      </div>
      {% endif %}

      {% else %}
      <div class="text-center py-5 text-muted">
        <i class="bi bi-calendar-x fs-1 opacity-25"></i>
        <p class="mt-3">لا توجد سجلات حضور للفترة المحددة</p>
      </div>
      {% endif %}
    </div>
  </div>

</div>
{% endblock %}
"""

write_file('templates/attendance/list.html', attendance_list_template)


# ════════════════════════════════════════════════
# FIX 2: reports/utils.py — field names صح
# ════════════════════════════════════════════════
print("\n📌 Fix 2: reports/utils.py — field names صح")

reports_utils = '''"""
reports/utils.py — Patch 49a Fix3
الـ field names الصح:
  check_in_time / check_out_time
  company.name_ar
  department.name_ar
"""

from datetime import date, timedelta, datetime
from django.http import HttpResponse
from django.utils import timezone

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ════════════════════════════════════
# Date Range
# ════════════════════════════════════

def get_date_range(period, custom_start=None, custom_end=None):
    today = date.today()
    if period == "week":
        start = today - timedelta(days=today.weekday())
        end   = start + timedelta(days=6)
    elif period == "month":
        start = today.replace(day=1)
        next_m = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
        end   = next_m - timedelta(days=1)
    elif period == "quarter":
        q     = (today.month - 1) // 3
        start = date(today.year, q * 3 + 1, 1)
        end   = (date(today.year, (q+1)*3+1, 1) - timedelta(days=1)) if q < 3 else date(today.year, 12, 31)
    elif period == "year":
        start = today.replace(month=1, day=1)
        end   = today.replace(month=12, day=31)
    elif period == "custom" and custom_start and custom_end:
        start, end = custom_start, custom_end
    else:
        start = today.replace(day=1)
        end   = today
    return start, end


# ════════════════════════════════════
# Attendance Summary
# ════════════════════════════════════

def get_attendance_summary(company, start_date, end_date):
    from attendance.models import Attendance
    qs      = Attendance.objects.filter(company=company, date__range=(start_date, end_date))
    total   = qs.count()
    present = qs.filter(status='present').count()
    late    = qs.filter(status='late').count()
    absent  = qs.filter(status='absent').count()
    on_leave= qs.filter(status='on_leave').count()
    return {
        'total':    total,
        'present':  present,
        'late':     late,
        'absent':   absent,
        'on_leave': on_leave,
        'rate':     round(present / total * 100, 1) if total else 0,
    }


# ════════════════════════════════════
# Employee Attendance Details
# ════════════════════════════════════

def get_employee_attendance_details(company, start_date, end_date):
    from attendance.models import Attendance
    from employees.models import Employee
    from django.db.models import Sum

    employees = Employee.objects.filter(company=company, status='active')
    result = []

    for emp in employees:
        qs      = Attendance.objects.filter(employee=emp, date__range=(start_date, end_date))
        total   = qs.count()
        present = qs.filter(status='present').count()
        late    = qs.filter(status='late').count()
        absent  = qs.filter(status='absent').count()
        on_leave= qs.filter(status='on_leave').count()
        late_mins = qs.aggregate(Sum('late_minutes'))['late_minutes__sum'] or 0

        # ساعات العمل من work_hours field
        from django.db.models import Sum as S
        work_hours_sum = qs.aggregate(S('work_hours'))['work_hours__sum']
        if work_hours_sum is None:
            # احسب يدوي من check_in_time و check_out_time
            work_seconds = 0
            for a in qs:
                if a.check_in_time and a.check_out_time:
                    ci = datetime.combine(date.today(), a.check_in_time)
                    co = datetime.combine(date.today(), a.check_out_time)
                    if co > ci:
                        work_seconds += (co - ci).seconds
            work_hours = round(work_seconds / 3600, 1)
        else:
            work_hours = round(float(work_hours_sum), 1)

        result.append({
            'employee':   emp,
            'total':      total,
            'present':    present,
            'late':       late,
            'absent':     absent,
            'on_leave':   on_leave,
            'rate':       round(present / total * 100, 1) if total else 0,
            'late_mins':  late_mins,
            'work_hours': work_hours,
        })
    return result


# ════════════════════════════════════
# Field Tracking Summary
# ════════════════════════════════════

def get_field_tracking_summary(company, start_date, end_date):
    from attendance.models import LocationLog
    from employees.models import Employee

    employees = Employee.objects.filter(company=company, status='active', is_field_worker=True)
    result = []
    for emp in employees:
        logs = LocationLog.objects.filter(
            employee=emp,
            timestamp__date__range=(start_date, end_date)
        )
        result.append({
            'employee':  emp,
            'log_count': logs.count(),
            'last_log':  logs.order_by('-timestamp').first(),
        })
    return result


# ════════════════════════════════════
# Excel Export — openpyxl احترافي
# ════════════════════════════════════

def export_to_excel(headers, rows, filename="report", company=None, report_title="تقرير"):
    if not OPENPYXL_AVAILABLE:
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f\'attachment; filename="{filename}.csv"\'
        import csv
        writer = csv.writer(response)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_title[:31]
    ws.sheet_view.rightToLeft = True

    PRIMARY   = "06B6D4"
    HEADER_BG = "0E7490"
    LIGHT_ROW = "F0FDFF"
    WHITE     = "FFFFFF"
    DARK      = "1E293B"
    GRAY      = "94A3B8"

    thin_border = Border(
        left=Side(style=\'thin\', color=\'DEE2E6\'),
        right=Side(style=\'thin\', color=\'DEE2E6\'),
        top=Side(style=\'thin\', color=\'DEE2E6\'),
        bottom=Side(style=\'thin\', color=\'DEE2E6\'),
    )

    num_cols = max(len(headers), 1)
    last_col = get_column_letter(num_cols)
    current_row = 1

    # ── اسم الشركة ──
    company_name = ""
    if company:
        company_name = getattr(company, \'name_ar\', None) or getattr(company, \'name_en\', None) or "MotionHR"
    else:
        company_name = "MotionHR"

    ws.merge_cells(f\'A{current_row}:{last_col}{current_row}\')
    cell = ws.cell(row=current_row, column=1, value=company_name)
    cell.font      = Font(name=\'Cairo\', size=16, bold=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal=\'center\', vertical=\'center\')
    ws.row_dimensions[current_row].height = 35
    current_row += 1

    # ── عنوان التقرير ──
    ws.merge_cells(f\'A{current_row}:{last_col}{current_row}\')
    cell = ws.cell(row=current_row, column=1, value=report_title)
    cell.font      = Font(name=\'Cairo\', size=13, bold=True, color=DARK)
    cell.alignment = Alignment(horizontal=\'center\', vertical=\'center\')
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # ── تاريخ التصدير ──
    ws.merge_cells(f\'A{current_row}:{last_col}{current_row}\')
    export_date = timezone.now().strftime("%Y/%m/%d %H:%M")
    cell = ws.cell(row=current_row, column=1, value=f"تاريخ التصدير: {export_date}")
    cell.font      = Font(name=\'Cairo\', size=10, color=GRAY)
    cell.alignment = Alignment(horizontal=\'center\', vertical=\'center\')
    ws.row_dimensions[current_row].height = 20
    current_row += 2

    # ── Header ──
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.font      = Font(name=\'Cairo\', size=11, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal=\'center\', vertical=\'center\', wrap_text=True)
        cell.border    = thin_border
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # ── Data ──
    for row_idx, row in enumerate(rows):
        bg = LIGHT_ROW if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name=\'Cairo\', size=10)
            cell.alignment = Alignment(horizontal=\'center\', vertical=\'center\', wrap_text=True)
            cell.border    = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # ── Footer ──
    current_row += 1
    ws.merge_cells(f\'A{current_row}:{last_col}{current_row}\')
    cell = ws.cell(row=current_row, column=1,
                   value="MotionHR — HR in Motion | JS Solution")
    cell.font      = Font(name=\'Cairo\', size=9, color=GRAY, italic=True)
    cell.alignment = Alignment(horizontal=\'center\', vertical=\'center\')

    # ── عرض الأعمدة ──
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for c in row:
                try:
                    if c.value:
                        max_len = max(max_len, len(str(c.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type=\'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\'
    )
    response[\'Content-Disposition\'] = f\'attachment; filename="{filename}.xlsx"\'
    wb.save(response)
    return response
'''

write_file('reports/utils.py', reports_utils)


# ════════════════════════════════════════════════
# FIX 3: reports/views.py — field names صح
# ════════════════════════════════════════════════
print("\n📌 Fix 3: reports/views.py — field names صح")

reports_views = '''"""
reports/views.py — Patch 49a Fix3
field names صح: name_ar, check_in_time, check_out_time
"""

from datetime import date
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from .utils import (
    get_date_range,
    get_attendance_summary,
    get_employee_attendance_details,
    get_field_tracking_summary,
    export_to_excel,
)


@login_required
def reports_home(request):
    return render(request, "reports/home.html", {"page_title": "التقارير"})


# ════════════════════════════════════
# تقرير الحضور
# ════════════════════════════════════

@login_required
def attendance_report(request):
    company = request.user.company
    period  = request.GET.get("period", "month")

    custom_start = custom_end = None
    if period == "custom":
        try:
            custom_start = date.fromisoformat(request.GET.get("start_date", ""))
            custom_end   = date.fromisoformat(request.GET.get("end_date", ""))
        except Exception:
            period = "month"

    start_date, end_date = get_date_range(period, custom_start, custom_end)
    summary = get_attendance_summary(company, start_date, end_date)
    details = get_employee_attendance_details(company, start_date, end_date)

    if request.GET.get("export") == "excel":
        headers = [
            "الموظف", "الرقم الوظيفي", "الحاضر", "الغائب",
            "المتأخر", "إجازة", "الإجمالي", "نسبة الحضور%",
            "إجمالي التأخير (دقيقة)", "ساعات العمل"
        ]
        rows = [[
            d["employee"].full_name_ar,
            d["employee"].employee_code,
            d["present"], d["absent"], d["late"],
            d["on_leave"], d["total"], d["rate"],
            d["late_mins"], d["work_hours"],
        ] for d in details]
        return export_to_excel(
            headers, rows,
            filename=f"attendance_{start_date}_{end_date}",
            company=company,
            report_title=f"تقرير الحضور والغياب — {start_date} إلى {end_date}"
        )

    return render(request, "reports/attendance_report.html", {
        "page_title": "تقرير الحضور والغياب",
        "summary": summary, "details": details,
        "start_date": start_date, "end_date": end_date, "period": period,
    })


# ════════════════════════════════════
# تقرير التأخيرات
# ════════════════════════════════════

@login_required
def late_report(request):
    from attendance.models import Attendance
    company = request.user.company
    period  = request.GET.get("period", "month")

    custom_start = custom_end = None
    if period == "custom":
        try:
            custom_start = date.fromisoformat(request.GET.get("start_date", ""))
            custom_end   = date.fromisoformat(request.GET.get("end_date", ""))
        except Exception:
            period = "month"

    start_date, end_date = get_date_range(period, custom_start, custom_end)

    late_records = Attendance.objects.filter(
        company=company,
        date__range=(start_date, end_date),
        status="late"
    ).select_related("employee", "shift").order_by("-date")

    if request.GET.get("export") == "excel":
        headers = ["الموظف", "الرقم الوظيفي", "التاريخ",
                   "وقت الدخول", "التأخير (دقيقة)", "الوردية"]
        rows = [[
            r.employee.full_name_ar,
            r.employee.employee_code,
            str(r.date),
            str(r.check_in_time) if r.check_in_time else "—",
            r.late_minutes or 0,
            r.shift.name if r.shift else "—",
        ] for r in late_records]
        return export_to_excel(
            headers, rows,
            filename=f"late_{start_date}_{end_date}",
            company=company,
            report_title=f"تقرير التأخيرات — {start_date} إلى {end_date}"
        )

    return render(request, "reports/late_report.html", {
        "page_title": "تقرير التأخيرات",
        "late_records": late_records,
        "start_date": start_date, "end_date": end_date, "period": period,
    })


# ════════════════════════════════════
# تقرير الإجازات
# ════════════════════════════════════

@login_required
def leave_report(request):
    from leaves.models import LeaveRequest
    company = request.user.company
    period  = request.GET.get("period", "month")

    custom_start = custom_end = None
    if period == "custom":
        try:
            custom_start = date.fromisoformat(request.GET.get("start_date", ""))
            custom_end   = date.fromisoformat(request.GET.get("end_date", ""))
        except Exception:
            period = "month"

    start_date, end_date = get_date_range(period, custom_start, custom_end)

    leave_requests = LeaveRequest.objects.filter(
        employee__company=company,
        start_date__range=(start_date, end_date)
    ).select_related("employee", "leave_type").order_by("-start_date")

    if request.GET.get("export") == "excel":
        headers = ["الموظف", "الرقم الوظيفي", "نوع الإجازة",
                   "من", "إلى", "الأيام", "الحالة"]
        rows = []
        for r in leave_requests:
            # حساب الأيام
            try:
                days = (r.end_date - r.start_date).days + 1
            except Exception:
                days = "—"
            # اسم نوع الإجازة
            leave_type_name = "—"
            if r.leave_type:
                leave_type_name = (
                    getattr(r.leave_type, 'name_ar', None) or
                    getattr(r.leave_type, 'name', None) or "—"
                )
            rows.append([
                r.employee.full_name_ar,
                r.employee.employee_code,
                leave_type_name,
                str(r.start_date),
                str(r.end_date),
                days,
                r.get_status_display() if hasattr(r, 'get_status_display') else r.status,
            ])
        return export_to_excel(
            headers, rows,
            filename=f"leaves_{start_date}_{end_date}",
            company=company,
            report_title=f"تقرير الإجازات — {start_date} إلى {end_date}"
        )

    return render(request, "reports/leave_report.html", {
        "page_title": "تقرير الإجازات",
        "leave_requests": leave_requests,
        "start_date": start_date, "end_date": end_date, "period": period,
    })


# ════════════════════════════════════
# تقرير الموظفين
# ════════════════════════════════════

@login_required
def employees_report(request):
    from employees.models import Employee
    company = request.user.company

    employees = Employee.objects.filter(
        company=company
    ).select_related("branch", "department", "job_title").order_by("employee_code")

    status_filter = request.GET.get("status", "")
    if status_filter:
        employees = employees.filter(status=status_filter)

    if request.GET.get("export") == "excel":
        headers = [
            "الرقم الوظيفي", "الاسم", "القسم", "الفرع",
            "المسمى الوظيفي", "تاريخ التعيين", "الحالة", "نوع العقد"
        ]
        rows = []
        for emp in employees:
            # Department name
            dept_name = "—"
            if emp.department:
                dept_name = (
                    getattr(emp.department, 'name_ar', None) or
                    getattr(emp.department, 'name_en', None) or "—"
                )
            # Branch name
            branch_name = "—"
            if emp.branch:
                branch_name = (
                    getattr(emp.branch, 'name_ar', None) or
                    getattr(emp.branch, 'name_en', None) or
                    getattr(emp.branch, 'name', None) or "—"
                )
            # Job title
            job_name = "—"
            if emp.job_title:
                job_name = (
                    getattr(emp.job_title, 'name_ar', None) or
                    getattr(emp.job_title, 'name_en', None) or "—"
                )
            rows.append([
                emp.employee_code,
                emp.full_name_ar,
                dept_name,
                branch_name,
                job_name,
                str(emp.hire_date) if emp.hire_date else "—",
                emp.get_status_display() if hasattr(emp, 'get_status_display') else emp.status,
                emp.get_contract_type_display() if hasattr(emp, 'get_contract_type_display') else emp.contract_type,
            ])
        return export_to_excel(
            headers, rows,
            filename=f"employees_{date.today()}",
            company=company,
            report_title="تقرير الموظفين"
        )

    return render(request, "reports/employees_report.html", {
        "page_title": "تقرير الموظفين",
        "employees": employees,
        "status_filter": status_filter,
    })


# ════════════════════════════════════
# تقرير الميدانيين
# ════════════════════════════════════

@login_required
def field_report(request):
    company = request.user.company
    period  = request.GET.get("period", "month")

    custom_start = custom_end = None
    if period == "custom":
        try:
            custom_start = date.fromisoformat(request.GET.get("start_date", ""))
            custom_end   = date.fromisoformat(request.GET.get("end_date", ""))
        except Exception:
            period = "month"

    start_date, end_date = get_date_range(period, custom_start, custom_end)
    tracking_data = get_field_tracking_summary(company, start_date, end_date)

    return render(request, "reports/field_report.html", {
        "page_title": "تقرير الموظفين الميدانيين",
        "tracking_data": tracking_data,
        "start_date": start_date, "end_date": end_date, "period": period,
    })
'''

write_file('reports/views.py', reports_views)


# ════════════════════════════════════════════════
# FIX 4: Stealth Policy — تفعيل مباشر من الـ view
# ════════════════════════════════════════════════
print("\n📌 Fix 4: تفعيل stealth_tracking من الـ DB مباشرة")

try:
    from companies.models import CompanyPolicy
    from accounts.models import User

    admin_user = User.objects.filter(role='company_admin').first()
    if admin_user and admin_user.company:
        policy, created = CompanyPolicy.objects.get_or_create(
            company=admin_user.company
        )
        if not policy.stealth_tracking_enabled:
            policy.stealth_tracking_enabled = True
            policy.save()
            print("   ✅ تم تفعيل stealth_tracking_enabled في قاعدة البيانات")
        else:
            print("   ℹ️ stealth_tracking_enabled مفعّل بالفعل")
    else:
        print("   ⚠️ لا يوجد company_admin — فعّله يدوياً من سياسات الشركة")
except Exception as e:
    print(f"   ⚠️ خطأ: {e}")


# ════════════════════════════════════════════════
# FIX 5: visit form — fix dashboard URL
# ════════════════════════════════════════════════
print("\n📌 Fix 5: Visit Form — fix dashboard URL")

visit_content = read_file('templates/attendance/visit_form.html')
if visit_content and 'dashboard:index' in visit_content:
    visit_content = visit_content.replace(
        "{% url 'dashboard:index' %}",
        "{% url 'dashboard' %}"
    )
    write_file('templates/attendance/visit_form.html', visit_content)
    print("   ✅ تم تصحيح dashboard URL في visit_form")
else:
    print("   ℹ️ visit_form — dashboard URL تمام")


print("\n" + "=" * 60)
print("✅ Patch 49a Fix3 اكتمل!")
print("=" * 60)
print("""
الإصلاحات:
  ✅ attendance/list.html   → dashboard URL + check_in_time
  ✅ reports/utils.py       → name_ar + check_in_time + check_out_time
  ✅ reports/views.py       → name_ar لكل الـ models
  ✅ stealth_tracking       → تم تفعيله في قاعدة البيانات

شغّل:
  python manage.py check
  python manage.py runserver 0.0.0.0:8000
""")