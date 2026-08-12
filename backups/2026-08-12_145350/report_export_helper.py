"""
Report Export Helper - نظام موحد لتصدير التقارير
كل التقارير بتستخدم النظام ده للـ Excel و PDF مع لوجو الشركة
"""
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone


def get_company_info(user):
    """جلب بيانات الشركة"""
    company = getattr(user, 'company', None)
    if not company:
        return {}

    return {
        'name_ar': getattr(company, 'name_ar', '') or getattr(company, 'name', ''),
        'name_en': getattr(company, 'name_en', '') or '',
        'logo_url': company.logo.url if getattr(company, 'logo', None) else None,
        'logo_path': company.logo.path if getattr(company, 'logo', None) else None,
        'phone': getattr(company, 'phone', '') or '',
        'email': getattr(company, 'email', '') or '',
        'address': getattr(company, 'address', '') or '',
    }


def export_to_excel(title, columns, rows, user, filename=None, subtitle=None):
    """
    تصدير تقرير كـ Excel مع لوجو الشركة

    Args:
        title: عنوان التقرير
        columns: قائمة [(key, label, width)]
        rows: قائمة dict فيها البيانات
        user: المستخدم
        filename: اسم الملف (اختياري)
        subtitle: عنوان فرعي (اختياري)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    company = get_company_info(user)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limit
    ws.sheet_view.rightToLeft = True

    # صف اللوجو والاسم
    row = 1
    logo_added = False
    if company.get('logo_path'):
        try:
            img = XLImage(company['logo_path'])
            img.width = 80
            img.height = 80
            ws.add_image(img, f'A{row}')
            ws.row_dimensions[row].height = 65
            ws.row_dimensions[row + 1].height = 20
            logo_added = True
        except Exception:
            pass

    # اسم الشركة
    name_col = 'C' if logo_added else 'A'
    ws[f'{name_col}{row}'] = company.get('name_ar', '') or company.get('name_en', '') or 'MotionHR'
    ws[f'{name_col}{row}'].font = Font(size=18, bold=True, color='0891B2')
    ws.merge_cells(f'{name_col}{row}:{get_column_letter(len(columns))}{row}')

    # معلومات الاتصال
    info_parts = []
    if company.get('phone'):
        info_parts.append(f"📞 {company['phone']}")
    if company.get('email'):
        info_parts.append(f"📧 {company['email']}")
    if info_parts:
        ws[f'{name_col}{row+1}'] = ' | '.join(info_parts)
        ws[f'{name_col}{row+1}'].font = Font(size=10, color='6B7280')
        ws.merge_cells(f'{name_col}{row+1}:{get_column_letter(len(columns))}{row+1}')

    # عنوان التقرير
    row += 3
    ws.cell(row=row, column=1, value=title).font = Font(size=16, bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

    # subtitle
    if subtitle:
        row += 1
        ws.cell(row=row, column=1, value=subtitle).font = Font(size=11, italic=True, color='6B7280')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

    # تاريخ التصدير
    row += 1
    export_date = timezone.now().strftime('%Y-%m-%d %H:%M')
    ws.cell(row=row, column=1, value=f'تاريخ التصدير: {export_date}').font = Font(size=9, color='9CA3AF')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

    # صف الأعمدة
    row += 2
    header_fill = PatternFill(start_color='0891B2', end_color='0891B2', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    for idx, col in enumerate(columns, start=1):
        key, label = col[0], col[1]
        width = col[2] if len(col) > 2 else 20
        cell = ws.cell(row=row, column=idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[row].height = 30

    # البيانات
    for r_idx, row_data in enumerate(rows, start=row + 1):
        for c_idx, col in enumerate(columns, start=1):
            key = col[0]
            value = row_data.get(key, '') if isinstance(row_data, dict) else ''
            if value is None:
                value = ''
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

    # حفظ
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    if not filename:
        filename = f'{title}_{timezone.now().strftime("%Y%m%d")}.xlsx'

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_to_pdf(title, columns, rows, user, filename=None, subtitle=None):
    """
    تصدير تقرير كـ PDF مع لوجو الشركة
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, Image as RLImage
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import os

    company = get_company_info(user)

    # نحاول نسجل خط عربي
    arabic_font = 'Helvetica'
    try:
        arabic_font_path = '/var/www/motionhr/static/fonts/Cairo-Regular.ttf'
        if os.path.exists(arabic_font_path):
            pdfmetrics.registerFont(TTFont('Cairo', arabic_font_path))
            arabic_font = 'Cairo'
    except Exception:
        pass

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=15*mm, leftMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )

    story = []
    styles = getSampleStyleSheet()

    # اللوجو
    if company.get('logo_path'):
        try:
            logo = RLImage(company['logo_path'], width=50*mm, height=50*mm)
            story.append(logo)
            story.append(Spacer(1, 5*mm))
        except Exception:
            pass

    # اسم الشركة
    company_name = company.get('name_ar') or company.get('name_en') or 'MotionHR'
    company_style = ParagraphStyle(
        'CompanyName', parent=styles['Title'],
        fontName=arabic_font, fontSize=18,
        textColor=colors.HexColor('#0891B2'),
        alignment=TA_CENTER,
    )
    story.append(Paragraph(company_name, company_style))

    # معلومات الاتصال
    info_parts = []
    if company.get('phone'):
        info_parts.append(company['phone'])
    if company.get('email'):
        info_parts.append(company['email'])

    if info_parts:
        info_style = ParagraphStyle(
            'Info', parent=styles['Normal'],
            fontName=arabic_font, fontSize=9,
            textColor=colors.HexColor('#6B7280'),
            alignment=TA_CENTER,
        )
        story.append(Paragraph(' | '.join(info_parts), info_style))

    story.append(Spacer(1, 5*mm))

    # عنوان التقرير
    title_style = ParagraphStyle(
        'Title', parent=styles['Heading1'],
        fontName=arabic_font, fontSize=16,
        alignment=TA_CENTER,
    )
    story.append(Paragraph(title, title_style))

    if subtitle:
        subtitle_style = ParagraphStyle(
            'Subtitle', parent=styles['Normal'],
            fontName=arabic_font, fontSize=11,
            textColor=colors.HexColor('#6B7280'),
            alignment=TA_CENTER,
        )
        story.append(Paragraph(subtitle, subtitle_style))

    export_date = timezone.now().strftime('%Y-%m-%d %H:%M')
    date_style = ParagraphStyle(
        'Date', parent=styles['Normal'],
        fontName=arabic_font, fontSize=9,
        textColor=colors.HexColor('#9CA3AF'),
        alignment=TA_CENTER,
    )
    story.append(Paragraph(f'Export Date: {export_date}', date_style))
    story.append(Spacer(1, 5*mm))

    # الجدول
    table_data = [[col[1] for col in columns]]
    for row_data in rows:
        table_row = []
        for col in columns:
            val = row_data.get(col[0], '') if isinstance(row_data, dict) else ''
            table_row.append(str(val) if val is not None else '')
        table_data.append(table_row)

    if len(table_data) > 1:
        col_widths = [(doc.width / len(columns))] * len(columns)
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0891B2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(table)
    else:
        story.append(Paragraph('لا توجد بيانات', styles['Normal']))

    doc.build(story)
    buffer.seek(0)

    if not filename:
        filename = f'{title}_{timezone.now().strftime("%Y%m%d")}.pdf'

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
