"""
Report Export Helper - نظام موحد لتصدير التقارير
كل التقارير بتستخدم النظام ده للـ Excel و PDF مع لوجو الشركة
"""
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone


def get_company_info(user):
    """جلب بيانات الشركة"""
    company = getattr(user, 'company', None)
    if not company:
        return {}

    return {
        'name_ar': getattr(company, 'name_ar', '') or getattr(company, 'name', ''),
        'name_en': getattr(company, 'name_en', '') or '',
        'logo_url': company.logo.url if getattr(company, 'logo', None) else None,
        'logo_path': company.logo.path if getattr(company, 'logo', None) else None,
        'phone': getattr(company, 'phone', '') or '',
        'email': getattr(company, 'email', '') or '',
        'address': getattr(company, 'address', '') or '',
    }


def export_to_excel(title, columns, rows, user, filename=None, subtitle=None):
    """
    تصدير تقرير كـ Excel مع لوجو الشركة

    Args:
        title: عنوان التقرير
        columns: قائمة [(key, label, width)]
        rows: قائمة dict فيها البيانات
        user: المستخدم
        filename: اسم الملف (اختياري)
        subtitle: عنوان فرعي (اختياري)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    # عكس ترتيب الأعمدة عشان يبقى متوافق مع الاتجاه العربي RTL (أول عمود يظهر يمين)
    columns = list(reversed(columns))

    company = get_company_info(user)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limit
    ws.sheet_view.rightToLeft = True

    # اللوجو + اسم الشركة (يمين لأن الشيت RTL)
    row = 1
    logo_added = False
    last_col_letter = get_column_letter(len(columns))
    if company.get('logo_path'):
        try:
            img = XLImage(company['logo_path'])
            img.width = 45
            img.height = 45
            ws.add_image(img, f'{last_col_letter}{row}')
            ws.row_dimensions[row].height = 35
            logo_added = True
        except Exception:
            pass

    # اسم الشركة (تحت اللوجو، نفس عمود اللوجو، وسط الصفحة لباقي الأعمدة)
    name_row = row + (3 if logo_added else 0)
    ws.cell(row=name_row, column=1, value=company.get('name_ar', '') or company.get('name_en', '') or 'MotionHR')
    ws.cell(row=name_row, column=1).font = Font(size=16, bold=True, color='0891B2')
    ws.cell(row=name_row, column=1).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=name_row, start_column=1, end_row=name_row, end_column=len(columns))

    # معلومات الاتصال
    info_parts = []
    if company.get('phone'):
        info_parts.append(f"📞 {company['phone']}")
    if company.get('email'):
        info_parts.append(f"📧 {company['email']}")
    if info_parts:
        info_row = name_row + 1
        ws.cell(row=info_row, column=1, value=' | '.join(info_parts))
        ws.cell(row=info_row, column=1).font = Font(size=10, color='6B7280')
        ws.cell(row=info_row, column=1).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=len(columns))
        row = info_row
    else:
        row = name_row

    # عنوان التقرير
    row += 3
    ws.cell(row=row, column=1, value=title).font = Font(size=16, bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

    # subtitle
    if subtitle:
        row += 1
        ws.cell(row=row, column=1, value=subtitle).font = Font(size=11, italic=True, color='6B7280')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

    # تاريخ التصدير
    row += 1
    export_date = timezone.now().strftime('%Y-%m-%d %H:%M')
    ws.cell(row=row, column=1, value=f'تاريخ التصدير: {export_date}').font = Font(size=9, color='9CA3AF')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

    # صف الأعمدة
    row += 2
    header_fill = PatternFill(start_color='0891B2', end_color='0891B2', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    for idx, col in enumerate(columns, start=1):
        key, label = col[0], col[1]
        width = col[2] if len(col) > 2 else 20
        cell = ws.cell(row=row, column=idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[row].height = 30

    # البيانات
    for r_idx, row_data in enumerate(rows, start=row + 1):
        for c_idx, col in enumerate(columns, start=1):
            key = col[0]
            value = row_data.get(key, '') if isinstance(row_data, dict) else ''
            if value is None:
                value = ''
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

    # حفظ
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    if not filename:
        filename = f'{title}_{timezone.now().strftime("%Y%m%d")}.xlsx'

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_to_pdf(title, columns, rows, user, filename=None, subtitle=None):
    """
    تصدير تقرير كـ PDF مع لوجو الشركة
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, Image as RLImage
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import os
    import re
    import arabic_reshaper
    from bidi.algorithm import get_display

    company = get_company_info(user)

    # عكس ترتيب الأعمدة عشان يبقى متوافق مع الاتجاه العربي RTL (أول عمود يظهر يمين)
    columns = list(reversed(columns))

    def shape_rtl(value):
        if value is None:
            return ''
        text = str(value)
        if not text:
            return ''
        if re.search(r'[؀-ۿ]', text):
            try:
                return get_display(arabic_reshaper.reshape(text))
            except Exception:
                return text
        return text

    # نحاول نسجل خط عربي من المسارات المتاحة فعلاً
    arabic_font = 'Helvetica'
    arabic_font_candidates = [
        ('Amiri', '/var/www/motionhr/core/fonts/Amiri-Regular.ttf'),
        ('Amiri', '/usr/share/fonts/opentype/fonts-hosny-amiri/Amiri-Regular.ttf'),
        ('Cairo', '/var/www/motionhr/core/fonts/Cairo-Regular.ttf'),
        ('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'),
    ]
    try:
        registered = set(pdfmetrics.getRegisteredFontNames())
        for font_name, font_path in arabic_font_candidates:
            if os.path.exists(font_path):
                if font_name not in registered:
                    pdfmetrics.registerFont(TTFont(font_name, font_path))
                arabic_font = font_name
                break
    except Exception:
        pass

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=15*mm, leftMargin=15*mm,
        topMargin=45*mm, bottomMargin=15*mm,
    )

    styles = getSampleStyleSheet()
    company_name = company.get('name_ar') or company.get('name_en') or 'MotionHR'

    def draw_header(canvas, doc_obj):
        canvas.saveState()
        page_width = doc_obj.pagesize[0]
        top_y = doc_obj.pagesize[1] - 15*mm

        # اللوجو (يمين)
        logo_w, logo_h = 20*mm, 20*mm
        logo_x = page_width - 15*mm - logo_w
        logo_y = top_y - logo_h
        if company.get('logo_path'):
            try:
                canvas.drawImage(
                    company['logo_path'], logo_x, logo_y,
                    width=logo_w, height=logo_h,
                    preserveAspectRatio=True, mask='auto',
                )
            except Exception:
                pass

        # اسم الشركة تحت اللوجو
        canvas.setFont(arabic_font, 11)
        canvas.setFillColor(colors.HexColor('#0891B2'))
        canvas.drawCentredString(logo_x + logo_w / 2, logo_y - 5*mm, shape_rtl(company_name))

        info_parts = []
        if company.get('phone'):
            info_parts.append(company['phone'])
        if company.get('email'):
            info_parts.append(company['email'])
        if info_parts:
            canvas.setFont(arabic_font, 7)
            canvas.setFillColor(colors.HexColor('#6B7280'))
            canvas.drawCentredString(logo_x + logo_w / 2, logo_y - 9*mm, shape_rtl(' | '.join(info_parts)))

        # عنوان التقرير في نص الصفحة
        canvas.setFont(arabic_font, 15)
        canvas.setFillColor(colors.black)
        canvas.drawCentredString(page_width / 2, top_y - 7*mm, shape_rtl(title))

        if subtitle:
            canvas.setFont(arabic_font, 10)
            canvas.setFillColor(colors.HexColor('#6B7280'))
            canvas.drawCentredString(page_width / 2, top_y - 13*mm, shape_rtl(subtitle))

        export_date = timezone.now().strftime('%Y-%m-%d %H:%M')
        canvas.setFont(arabic_font, 8)
        canvas.setFillColor(colors.HexColor('#9CA3AF'))
        canvas.drawCentredString(page_width / 2, top_y - 18*mm, f'Export Date: {export_date}')

        canvas.restoreState()

    story = []

    # الجدول
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=10,
        leading=12,
        alignment=TA_RIGHT,
        textColor=colors.white,
        wordWrap='RTL',
    )
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=9,
        leading=11,
        alignment=TA_RIGHT,
        textColor=colors.black,
        wordWrap='RTL',
    )

    table_data = [[Paragraph(shape_rtl(col[1]), table_header_style) for col in columns]]
    for row_data in rows:
        table_row = []
        for col in columns:
            val = row_data.get(col[0], '') if isinstance(row_data, dict) else ''
            table_row.append(Paragraph(shape_rtl(val), table_cell_style))
        table_data.append(table_row)

    if len(table_data) > 1:
        col_widths = [(doc.width / len(columns))] * len(columns)
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0891B2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(table)
    else:
        no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontName=arabic_font, fontSize=11, alignment=TA_CENTER)
        story.append(Paragraph(shape_rtl('لا توجد بيانات'), no_data_style))

    doc.build(story, onFirstPage=draw_header, onLaterPages=draw_header)
    buffer.seek(0)

    if not filename:
        filename = f'{title}_{timezone.now().strftime("%Y%m%d")}.pdf'

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
