"""
employees/exports.py — Patch 49b Fix8
Excel + PDF export للموظفين — مع دعم عربي حقيقي في PDF
"""

import os
import re
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone

# ── Excel ───────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── PDF Arabic stack ────────────────────────────────────────
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import arabic_reshaper
    HAS_ARABIC_RESHAPER = True
except ImportError:
    HAS_ARABIC_RESHAPER = False

try:
    from bidi.algorithm import get_display
    HAS_BIDI = True
except ImportError:
    HAS_BIDI = False

from xml.sax.saxutils import escape


# ═════════════════════════════════════════════════════════════
# Shared helpers
# ═════════════════════════════════════════════════════════════

def get_company_name(user):
    """جلب اسم الشركة من المستخدم"""
    company = getattr(user, 'company', None)
    if company:
        return getattr(company, 'name_ar', None) or getattr(company, 'name_en', None) or 'MotionHR'
    return 'MotionHR'


def _safe_text(value, default='—'):
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def _employee_full_name_ar(emp):
    return (
        getattr(emp, 'full_name_ar', None)
        or f"{getattr(emp, 'first_name_ar', '')} {getattr(emp, 'last_name_ar', '')}".strip()
        or '—'
    )


def get_employee_row(emp):
    """تحويل موظف إلى صف بيانات"""
    dept = getattr(emp, 'department', None)
    branch = getattr(emp, 'branch', None)
    job = getattr(emp, 'job_title', None)
    manager = getattr(emp, 'direct_manager', None)

    return [
        _safe_text(getattr(emp, 'employee_code', None)),
        _safe_text(_employee_full_name_ar(emp)),
        _safe_text(f"{getattr(emp, 'first_name_en', '')} {getattr(emp, 'last_name_en', '')}".strip()),
        _safe_text(getattr(emp, 'national_id', None)),
        _safe_text((getattr(job, 'name_ar', None) or getattr(job, 'name_en', None)) if job else None),
        _safe_text((getattr(dept, 'name_ar', None) or getattr(dept, 'name_en', None)) if dept else None),
        _safe_text((getattr(branch, 'name_ar', None) or getattr(branch, 'name_en', None)) if branch else None),
        _safe_text(_employee_full_name_ar(manager) if manager else None),
        _safe_text(getattr(emp, 'phone', None)),
        _safe_text(getattr(emp, 'email', None)),
        _safe_text(getattr(emp, 'hire_date', None)),
        _safe_text(emp.get_contract_type_display() if hasattr(emp, 'get_contract_type_display') else getattr(emp, 'contract_type', None)),
        _safe_text(emp.get_status_display() if hasattr(emp, 'get_status_display') else getattr(emp, 'status', None)),
        _safe_text(getattr(emp, 'basic_salary', None)),
        _safe_text(emp.get_gender_display() if hasattr(emp, 'get_gender_display') else getattr(emp, 'gender', None)),
        _safe_text(getattr(emp, 'city', None)),
    ]


HEADERS = [
    "الرقم الوظيفي",
    "الاسم بالعربي",
    "الاسم بالإنجليزي",
    "الرقم القومي",
    "المسمى الوظيفي",
    "الإدارة",
    "الفرع",
    "المدير المباشر",
    "التليفون",
    "الإيميل",
    "تاريخ التعيين",
    "نوع العقد",
    "الحالة",
    "الراتب الأساسي",
    "النوع",
    "المدينة",
]


# ═════════════════════════════════════════════════════════════
# Excel Export
# ═════════════════════════════════════════════════════════════

def export_employees_excel(employees, user=None):
    """تصدير Excel احترافي"""
    if not HAS_OPENPYXL:
        return _fallback_csv(employees)

    company_name = get_company_name(user) if user else 'MotionHR'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الموظفون"
    ws.sheet_view.rightToLeft = True

    PRIMARY   = "06B6D4"
    HEADER_BG = "0E7490"
    LIGHT_ROW = "F0FDFF"
    WHITE     = "FFFFFF"
    DARK      = "1E293B"
    GRAY      = "94A3B8"

    thin_border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6'),
    )

    num_cols = len(HEADERS)
    last_col = get_column_letter(num_cols)
    row_num = 1

    ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
    cell = ws.cell(row=row_num, column=1, value=company_name)
    cell.font = Font(name='Cairo', size=16, bold=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_num].height = 35
    row_num += 1

    ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
    cell = ws.cell(row=row_num, column=1, value="تقرير الموظفين")
    cell.font = Font(name='Cairo', size=13, bold=True, color=DARK)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_num].height = 28
    row_num += 1

    ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
    export_date = timezone.now().strftime("%Y/%m/%d %H:%M")
    cell = ws.cell(row=row_num, column=1, value=f"تاريخ التصدير: {export_date}")
    cell.font = Font(name='Cairo', size=10, color=GRAY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_num].height = 20
    row_num += 2

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.font = Font(name='Cairo', size=11, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row_num].height = 30
    row_num += 1

    for idx, emp in enumerate(employees):
        bg = LIGHT_ROW if idx % 2 == 0 else WHITE
        row_data = get_employee_row(emp)
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name='Cairo', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[row_num].height = 22
        row_num += 1

    row_num += 1
    ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
    cell = ws.cell(row=row_num, column=1, value="MotionHR — HR in Motion | JS Solutions")
    cell.font = Font(name='Cairo', size=9, color=GRAY, italic=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(HEADERS[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for c in row:
                try:
                    if c.value:
                        max_len = max(max_len, len(str(c.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    today = timezone.now().strftime("%Y-%m-%d")
    response['Content-Disposition'] = f'attachment; filename="employees_{today}.xlsx"'
    wb.save(response)
    return response


def _fallback_csv(employees):
    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="employees.csv"'
    writer = csv.writer(response)
    writer.writerow(HEADERS)
    for emp in employees:
        writer.writerow(get_employee_row(emp))
    return response


# ═════════════════════════════════════════════════════════════
# Arabic PDF helpers
# ═════════════════════════════════════════════════════════════

_ARABIC_FONT_NAME = "MotionHRArabic"
_ARABIC_FONT_REGISTERED = False


def _contains_arabic(text):
    if not text:
        return False
    return bool(re.search(r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF]', str(text)))


def _shape_arabic_text(text):
    text = _safe_text(text, default='')
    if not text:
        return ''
    if not _contains_arabic(text):
        return text

    if HAS_ARABIC_RESHAPER:
        try:
            text = arabic_reshaper.reshape(text)
        except Exception:
            pass

    if HAS_BIDI:
        try:
            text = get_display(text)
        except Exception:
            pass

    return text


def _find_windows_arabic_font():
    candidates = [
        os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "tahoma.ttf"),
        os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "arial.ttf"),
        os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "trado.ttf"),
        os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "times.ttf"),
        os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "calibri.ttf"),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def _ensure_arabic_font_registered():
    global _ARABIC_FONT_REGISTERED

    if _ARABIC_FONT_REGISTERED:
        return _ARABIC_FONT_NAME

    if not HAS_REPORTLAB:
        return "Helvetica"

    font_path = _find_windows_arabic_font()
    if not font_path:
        return "Helvetica"

    try:
        pdfmetrics.registerFont(TTFont(_ARABIC_FONT_NAME, font_path))
        _ARABIC_FONT_REGISTERED = True
        return _ARABIC_FONT_NAME
    except Exception:
        return "Helvetica"


def _p(text, style):
    return Paragraph(escape(_shape_arabic_text(text)), style)


# ═════════════════════════════════════════════════════════════
# PDF Export — ReportLab + Arabic shaping
# ═════════════════════════════════════════════════════════════

def export_employees_pdf(employees, user=None):
    """تصدير PDF احترافي بعربي صحيح"""
    if not HAS_REPORTLAB:
        raise Exception("reportlab غير مثبت")

    font_name = _ensure_arabic_font_registered()
    company_name = get_company_name(user) if user else 'MotionHR'
    export_date = timezone.now().strftime("%Y/%m/%d %H:%M")
    today = timezone.now().strftime("%Y-%m-%d")

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=18,
        leftMargin=18,
        topMargin=18,
        bottomMargin=18,
        title="تقرير الموظفين",
        author="MotionHR",
    )

    title_style = ParagraphStyle(
        name="MotionHRTitle",
        fontName=font_name,
        fontSize=15,
        leading=18,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#06B6D4"),
    )

    subtitle_style = ParagraphStyle(
        name="MotionHRSubtitle",
        fontName=font_name,
        fontSize=11,
        leading=14,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#1E293B"),
    )

    meta_style = ParagraphStyle(
        name="MotionHRMeta",
        fontName=font_name,
        fontSize=8,
        leading=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#64748B"),
    )

    footer_style = ParagraphStyle(
        name="MotionHRFooter",
        fontName=font_name,
        fontSize=8,
        leading=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#94A3B8"),
    )

    story = []
    story.append(_p(company_name, title_style))
    story.append(Spacer(1, 4))
    story.append(_p("تقرير الموظفين", subtitle_style))
    story.append(Spacer(1, 3))
    story.append(_p(f"تاريخ التصدير: {export_date}", meta_style))
    story.append(Spacer(1, 10))

    headers = [_shape_arabic_text(h) for h in HEADERS]
    rows = [[_shape_arabic_text(v) for v in get_employee_row(emp)] for emp in employees]

    data = [headers] + rows

    col_widths = [
        40, 55, 50, 55, 50, 45, 45, 55,
        50, 70, 45, 45, 40, 50, 30, 40
    ]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0E7490")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('LEADING', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor("#D1D5DB")),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
        ('TOPPADDING', (0, 0), (-1, 0), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
    ])

    for row_idx in range(1, len(data)):
        bg = colors.HexColor("#F0FDFF") if row_idx % 2 == 1 else colors.white
        style.add('BACKGROUND', (0, row_idx), (-1, row_idx), bg)

    table.setStyle(style)
    story.append(table)
    story.append(Spacer(1, 10))
    story.append(_p("MotionHR — HR in Motion | JS Solutions", footer_style))

    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="employees_{today}.pdf"'
    response.write(pdf)
    return response
