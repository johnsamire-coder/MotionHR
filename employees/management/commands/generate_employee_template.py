"""
Management Command: generate_employee_template
ينشئ شيت إكسيل احترافي لاستيراد الموظفين - Version 2
خريطة الأعمدة ثابتة ومحددة ومتطابقة مع import_employees_bulk
"""
from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


# ═══════════════════════════════════════════════════════════════
# خريطة الأعمدة الثابتة — أي تعديل هنا لازم يتعمل في الاستيراد
# ═══════════════════════════════════════════════════════════════
COLUMNS = [
    # (key, label, required, val_type, val_formula, width, is_text_format, section)
    # --- العملية ---
    ("operation_type",            "نوع العملية *",                   True,  "list", "new,update",             15, False, "بيانات العملية"),
    ("employee_code",             "كود الموظف",                      False, None,   None,                     15, True,  "بيانات العملية"),
    ("temporary_password",        "الباسورد المؤقتة *",              True,  None,   None,                     20, True,  "بيانات العملية"),
    # --- الشخصية ---
    ("first_name_ar",             "الاسم الأول عربي *",              True,  None,   None,                     20, False, "البيانات الشخصية"),
    ("middle_name_ar",            "الاسم الأوسط عربي",               False, None,   None,                     20, False, "البيانات الشخصية"),
    ("last_name_ar",              "الاسم الأخير عربي *",             True,  None,   None,                     20, False, "البيانات الشخصية"),
    ("first_name_en",             "الاسم الأول إنجليزي *",           True,  None,   None,                     20, False, "البيانات الشخصية"),
    ("last_name_en",              "الاسم الأخير إنجليزي *",          True,  None,   None,                     20, False, "البيانات الشخصية"),
    ("national_id",               "الرقم القومي * (Text)",           True,  None,   None,                     20, True,  "البيانات الشخصية"),
    ("birth_date",                "تاريخ الميلاد * (YYYY-MM-DD)",    True,  None,   None,                     22, False, "البيانات الشخصية"),
    ("gender",                    "النوع *",                         True,  "list", "القوائم!$B$2:$B$3",      12, False, "البيانات الشخصية"),
    ("marital_status",            "الحالة الاجتماعية",               False, "list", "القوائم!$C$2:$C$5",      18, False, "البيانات الشخصية"),
    ("religion",                  "الديانة",                         False, "list", "القوائم!$D$2:$D$4",      12, False, "البيانات الشخصية"),
    ("nationality",               "الجنسية",                         False, None,   None,                     15, False, "البيانات الشخصية"),
    ("language",                  "اللغة",                           False, "list", "القوائم!$E$2:$E$3",      10, False, "البيانات الشخصية"),
    # --- التواصل ---
    ("country_code",              "كود الدولة",                      False, None,   None,                     12, False, "التواصل"),
    ("phone",                     "الموبايل * (Text)",               True,  None,   None,                     18, True,  "التواصل"),
    ("phone2",                    "موبايل إضافي (Text)",             False, None,   None,                     18, True,  "التواصل"),
    ("email",                     "البريد الإلكتروني",               False, None,   None,                     25, False, "التواصل"),
    ("address",                   "العنوان",                         False, None,   None,                     30, False, "التواصل"),
    ("city",                      "المدينة",                         False, None,   None,                     15, False, "التواصل"),
    ("emergency_contact_name",    "اسم جهة الطوارئ",                 False, None,   None,                     22, False, "التواصل"),
    ("emergency_contact_relation","صلة القرابة",                     False, None,   None,                     15, False, "التواصل"),
    ("emergency_contact_phone",   "موبايل الطوارئ (Text)",           False, None,   None,                     20, True,  "التواصل"),
    # --- الوظيفة ---
    ("branch_name", "الفرع *", True, "list", "التعريفات!$A$2:$A$200", 20, False, "بيانات الوظيفة"),
    ("department_name", "القسم *", True, "list", "التعريفات!$B$2:$B$200", 20, False, "بيانات الوظيفة"),
    ("job_title_name", "المسمى الوظيفي *", True, "list", "التعريفات!$C$2:$C$200", 22, False, "بيانات الوظيفة"),
    ("direct_manager_department", "قسم المدير المباشر", False, "list", "التعريفات!$B$2:$B$200", 22, False, "بيانات الوظيفة"),
    ("direct_manager_name", "اسم المدير المباشر", False, None, None, 28, False, "بيانات الوظيفة"),
    ("hire_date",                 "تاريخ التعيين * (YYYY-MM-DD)",    True,  None,   None,                     25, False, "بيانات الوظيفة"),
    ("attendance_mode",           "نمط الحضور *",                    True,  "list", "القوائم!$I$2:$I$6",      18, False, "بيانات الوظيفة"),
    ("status",                    "الحالة الوظيفية (اختياري - سيتم الاستيراد نشط)", False, "list", "القوائم!$O$2:$O$6", 28, False, "بيانات الوظيفة"),
    ("worker_type",               "تصنيف الموظف * (مكتبي/ميداني حر/ميداني محدد)", True, "list", "القوائم!$P$2:$P$4", 30, False, "بيانات الوظيفة"),
    # --- العقد ---
    ("contract_type",             "نوع العقد *",                     True,  "list", "القوائم!$J$2:$J$7",      18, False, "العقد"),
    ("contract_start_date",       "بداية العقد (YYYY-MM-DD)",        False, None,   None,                     22, False, "العقد"),
    ("contract_end_date",         "نهاية العقد (YYYY-MM-DD)",        False, None,   None,                     22, False, "العقد"),
    ("contract_duration_months",  "مدة العقد بالشهور (رقم فقط)",    False, None,   None,                     24, False, "العقد"),
    ("probation_months",          "فترة التجربة بالشهور (رقم فقط)", False, None,   None,                     24, False, "العقد"),
    # --- التأمين ---
    ("has_insurance",             "مؤمن عليه",                       False, "list", "نعم,لا",                 12, False, "التأمين"),
    ("insurance_number",          "رقم التأمين (Text)",              False, None,   None,                     18, True,  "التأمين"),
    # --- الماليات ---
    ("basic_salary",              "المرتب الأساسي",                  False, None,   None,                     18, False, "الماليات"),
    ("currency",                  "العملة",                          False, "list", "القوائم!$K$2:$K$7",      10, False, "الماليات"),
    ("salary_payment_method",     "طريقة القبض *",                   True,  "list", "القوائم!$L$2:$L$5",      18, False, "الماليات"),
    ("bank_name",                 "اسم البنك",                       False, None,   None,                     20, False, "الماليات"),
    ("bank_account",              "رقم الحساب (Text)",               False, None,   None,                     20, True,  "الماليات"),
    ("iban",                      "IBAN (Text - EG...)",             False, None,   None,                     30, True,  "الماليات"),
    ("instapay_transfer_id",      "رقم إنستا باي (Text)",            False, None,   None,                     22, True,  "الماليات"),
    ("wallet_transfer_number",    "رقم المحفظة (Text)",              False, None,   None,                     22, True,  "الماليات"),
    ("wallet_provider",           "مزود المحفظة",                    False, "list", "القوائم!$M$2:$M$7",      18, False, "الماليات"),
    # --- أرصدة الإجازات ---
    ("annual_entitled",           "السنوية - المستحق",               False, None,   None,                     20, False, "أرصدة الإجازات"),
    ("annual_used_before_system", "السنوية - المستنفذ",              False, None,   None,                     22, False, "أرصدة الإجازات"),
    ("annual_carry_forward",      "السنوية - مرحل",                  False, None,   None,                     18, False, "أرصدة الإجازات"),
    ("sick_entitled",             "المرضية - المستحق",               False, None,   None,                     20, False, "أرصدة الإجازات"),
    ("sick_used_before_system",   "المرضية - المستنفذ",              False, None,   None,                     22, False, "أرصدة الإجازات"),
    ("sick_carry_forward",        "المرضية - مرحل",                  False, None,   None,                     18, False, "أرصدة الإجازات"),
    ("emergency_entitled",        "الطارئة - المستحق",               False, None,   None,                     20, False, "أرصدة الإجازات"),
    ("emergency_used_before_system","الطارئة - المستنفذ",            False, None,   None,                     22, False, "أرصدة الإجازات"),
    ("emergency_carry_forward",   "الطارئة - مرحل",                  False, None,   None,                     18, False, "أرصدة الإجازات"),
    ("maternity_entitled",        "الأمومة - المستحق",               False, None,   None,                     20, False, "أرصدة الإجازات"),
    ("maternity_used_before_system","الأمومة - المستنفذ",            False, None,   None,                     22, False, "أرصدة الإجازات"),
    ("maternity_carry_forward",   "الأمومة - مرحل",                  False, None,   None,                     18, False, "أرصدة الإجازات"),
    ("paternity_entitled",        "الأبوة - المستحق",                False, None,   None,                     20, False, "أرصدة الإجازات"),
    ("paternity_used_before_system","الأبوة - المستنفذ",             False, None,   None,                     22, False, "أرصدة الإجازات"),
    ("paternity_carry_forward",   "الأبوة - مرحل",                   False, None,   None,                     18, False, "أرصدة الإجازات"),
    ("unpaid_entitled",           "بدون مرتب - المستحق",             False, None,   None,                     22, False, "أرصدة الإجازات"),
    ("unpaid_used_before_system", "بدون مرتب - المستنفذ",            False, None,   None,                     24, False, "أرصدة الإجازات"),
    ("unpaid_carry_forward",      "بدون مرتب - مرحل",                False, None,   None,                     20, False, "أرصدة الإجازات"),
]

# ألوان الأقسام
SECTION_COLORS = {
    "بيانات العملية":   "FFE082",
    "البيانات الشخصية": "CE93D8",
    "التواصل":          "80DEEA",
    "بيانات الوظيفة":  "90CAF9",
    "العقد":            "A5D6A7",
    "التأمين":          "FFCC80",
    "الماليات":         "F48FB1",
    "أرصدة الإجازات":  "B2DFDB",
}


class Command(BaseCommand):
    help = "ينشئ شيت Excel لاستيراد الموظفين - v2"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            type=str,
            default="/var/www/motionhr/media/employee_import_template.xlsx",
            help="مسار حفظ الملف",
        )

    def handle(self, *args, **options):
        output_path = options["output"]
        wb = Workbook()
        self._create_instructions_sheet(wb)
        self._create_employees_sheet(wb)
        self._create_definitions_sheet(wb)
        self._create_lists_sheet(wb)
        wb.save(output_path)
        self.stdout.write(self.style.SUCCESS(f"تم إنشاء الشيت بنجاح: {output_path}"))

    # ─────────────────────────────────────────
    def _create_instructions_sheet(self, wb):
        ws = wb.active
        ws.title = "تعليمات"

        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        bold_font   = Font(bold=True, size=11)
        red_font    = Font(color="C62828", bold=True)

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 65

        ws["A1"] = "نظام استيراد الموظفين - MotionHR v2"
        ws["A1"].fill = header_fill
        ws["A1"].font = header_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:B1")
        ws.row_dimensions[1].height = 35

        instructions = [
            ("", ""),
            ("قواعد عامة", ""),
            ("نوع العملية", "new = موظف جديد | update = تحديث موظف حالي"),
            ("الخانات بالنجمة (*)", "إجبارية — لو فاضية السيرفر هيرفض الصف ويقول السبب"),
            ("الخانات بدون نجمة", "اختيارية"),
            ("القوائم المنسدلة", "لازم تختار من القائمة فقط — الكتابة الحرة مش مقبولة"),
            ("", ""),
            ("تنسيقات مهمة", ""),
            ("الأعمدة النصية (Text)", "موبايل، رقم قومي، IBAN، باسبور، كود موظف، باسورد، رقم تأمين"),
            ("سبب Text في الموبايل", "عشان الصفر في الشمال ما يطيرش"),
            ("IBAN", "يبدأ بحروف زي EG — فلازم يبقى Text مش رقم"),
            ("مدة العقد / فترة التجربة", "أرقام فقط — يعني 12 تعني 12 شهر"),
            ("نهاية العقد", "لو اكتبت مدة العقد + بداية العقد — السيرفر هيحسب النهاية تلقائي"),
            ("", ""),
            ("المدير المباشر", ""),
            ("direct_manager_name", "اكتب اسم المدير كما هو في السيستم — السيرفر هيبحث ويربطه"),
            ("لو الاسم متكرر", "السيرفر هيرفض الصف ويطلب منك توضيح"),
            ("", ""),
            ("طريقة القبض", ""),
            ("cash", "مفيش حقول إضافية مطلوبة"),
            ("bank", "مطلوب: اسم البنك + رقم الحساب + اسم صاحب الحساب"),
            ("instapay", "مطلوب: رقم إنستا باي فقط"),
            ("wallet", "مطلوب: رقم المحفظة + مزود المحفظة"),
            ("", ""),
            ("التأمين", ""),
            ("لو مؤمن عليه = نعم", "رقم التأمين إجباري — السيرفر هيرفض لو فاضي"),
            ("لو مؤمن عليه = لا", "رقم التأمين مش مطلوب"),
            ("", ""),
            ("العقد", ""),
            ("permanent", "نهاية العقد مش إجبارية"),
            ("temporary / training / consultant", "بداية ونهاية العقد إجباريين"),
            ("", ""),
            ("أرصدة الإجازات", ""),
            ("entitled", "الرصيد المستحق — لو فاضي السيرفر يحسبه من السياسة تلقائي"),
            ("used_before_system", "الأيام اللي اتاخدت قبل تشغيل السيستم"),
            ("carry_forward", "الرصيد المرحل من سنة سابقة"),
            ("المتبقي", "السيستم بيحسبه = entitled + carry_forward - used"),
        ]

        for row_idx, (col_a, col_b) in enumerate(instructions, start=2):
            ws[f"A{row_idx}"] = col_a
            ws[f"B{row_idx}"] = col_b
            if col_b == "" and col_a != "":
                ws[f"A{row_idx}"].font = bold_font
                ws[f"A{row_idx}"].fill = PatternFill(
                    start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"
                )
                ws.merge_cells(f"A{row_idx}:B{row_idx}")
            elif "(*)" in col_a or "إجباري" in col_b:
                ws[f"A{row_idx}"].font = red_font

    # ─────────────────────────────────────────
    def _create_employees_sheet(self, wb):
        ws = wb.create_sheet("الموظفين")

        required_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        optional_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        header_font   = Font(bold=True, size=10)
        center        = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin          = Side(style="thin")
        border        = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Row 1 — Section headers
        current_section = None
        section_start   = 1
        for col_idx, col in enumerate(COLUMNS, start=1):
            section = col[7]
            if section != current_section:
                if current_section is not None:
                    end_col = col_idx - 1
                    cl_s = get_column_letter(section_start)
                    cl_e = get_column_letter(end_col)
                    ws.merge_cells(f"{cl_s}1:{cl_e}1")
                    cell = ws[f"{cl_s}1"]
                    cell.value = current_section
                    color = SECTION_COLORS.get(current_section, "EEEEEE")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = center
                    cell.border = border
                current_section = section
                section_start   = col_idx
        # آخر section
        end_col = len(COLUMNS)
        cl_s = get_column_letter(section_start)
        cl_e = get_column_letter(end_col)
        ws.merge_cells(f"{cl_s}1:{cl_e}1")
        cell = ws[f"{cl_s}1"]
        cell.value = current_section
        color = SECTION_COLORS.get(current_section, "EEEEEE")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True, size=10)
        cell.alignment = center
        cell.border = border

        # Row 2 — Column labels | Row 3 — Column keys
        for col_idx, (key, label, required, val_type, val_formula, width, is_text, section) in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

            # label
            cell = ws.cell(row=2, column=col_idx, value=label)
            cell.fill   = required_fill if required else optional_fill
            cell.font   = header_font
            cell.alignment = center
            cell.border = border

            # key
            key_cell = ws.cell(row=3, column=col_idx, value=key)
            key_cell.font      = Font(size=8, italic=True, color="757575")
            key_cell.alignment = center
            key_cell.border    = border

            # Data Validation مقفولة بـ showErrorMessage
            if val_type == "list" and val_formula:
                if val_formula.startswith("القوائم") or val_formula.startswith("التعريفات"):
                    formula = f"={val_formula}"
                else:
                    formula = f'"{val_formula}"'
                dv = DataValidation(
                    type="list",
                    formula1=formula,
                    allow_blank=not required,
                    showErrorMessage=True,
                    errorTitle="قيمة غير مسموح بها",
                    error="اختر قيمة من القائمة فقط",
                )
                dv.sqref = f"{col_letter}4:{col_letter}10000"
                ws.add_data_validation(dv)

            # Text format
            if is_text:
                for row in range(4, 10001):
                    ws.cell(row=row, column=col_idx).number_format = "@"

            # Number format لمدة العقد وفترة التجربة
            if key in ("contract_duration_months", "probation_months"):
                for row in range(4, 10001):
                    ws.cell(row=row, column=col_idx).number_format = "0"



        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 45
        ws.row_dimensions[3].height = 20
        ws.freeze_panes = "A4"

        # ─────────────────────────────────────────
        # Conditional Formatting — مطلوب / غير مطلوب
        # ─────────────────────────────────────────
        # بنحدد أعمدة الـ keys المهمة من COLUMNS
        col_map = {col[0]: idx for idx, col in enumerate(COLUMNS, start=1)}

        # طريقة القبض
        pay_col = get_column_letter(col_map.get("salary_payment_method", 0))

        # لون برتقالي للحقول المطلوبة حسب طريقة القبض
        orange_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
        # لون رمادي للحقول غير المطلوبة
        grey_fill   = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        bank_cols  = ["bank_name", "bank_account", "iban"]
        insta_cols = ["instapay_transfer_id"]
        wallet_cols = ["wallet_transfer_number", "wallet_provider"]

        for key_list, condition in [
            (bank_cols,   f'{pay_col}4="bank"'),
            (insta_cols,  f'{pay_col}4="instapay"'),
            (wallet_cols, f'{pay_col}4="wallet"'),
        ]:
            for key in key_list:
                c_idx = col_map.get(key)
                if not c_idx:
                    continue
                cl = get_column_letter(c_idx)
                rng = f"{cl}4:{cl}10000"
                # مطلوب = برتقالي
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(
                        formula=[f"${pay_col}4={condition.split(chr(61))[1]}"],
                        fill=orange_fill,
                    )
                )
                # غير مطلوب = رمادي
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(
                        formula=[f"NOT(${pay_col}4={condition.split(chr(61))[1]})"],
                        fill=grey_fill,
                    )
                )

        # التأمين
        ins_col = get_column_letter(col_map.get("has_insurance", 0))
        ins_num_col = col_map.get("insurance_number")
        if ins_num_col:
            cl = get_column_letter(ins_num_col)
            rng = f"{cl}4:{cl}10000"
            ws.conditional_formatting.add(
                rng,
                FormulaRule(
                    formula=[f'${ins_col}4="نعم"'],
                    fill=orange_fill,
                )
            )
            ws.conditional_formatting.add(
                rng,
                FormulaRule(
                    formula=[f'NOT(${ins_col}4="نعم")'],
                    fill=grey_fill,
                )
            )

    # ─────────────────────────────────────────
    def _create_definitions_sheet(self, wb):
        ws = wb.create_sheet('التعريفات')

        from companies.models import Branch, Department
        from employees.models import JobTitle

        header_fill = PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        center = Alignment(horizontal='center', vertical='center')

        headers = [
            ('A', 'الفروع الحالية / الجديدة'),
            ('B', 'الأقسام الحالية / الجديدة'),
            ('C', 'المسميات الوظيفية الحالية / الجديدة'),
        ]

        for col_letter, title in headers:
            ws.column_dimensions[col_letter].width = 35
            cell = ws[f'{col_letter}1']
            cell.value = title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        try:
            branches = list(Branch.objects.values_list('name_ar', flat=True).distinct())
        except Exception:
            branches = []

        try:
            departments = list(Department.objects.values_list('name_ar', flat=True).distinct())
        except Exception:
            departments = []

        try:
            job_titles = list(JobTitle.objects.values_list('name_ar', flat=True).distinct())
        except Exception:
            job_titles = []

        max_len = max(len(branches), len(departments), len(job_titles), 1)

        for i in range(max_len):
            ws.cell(row=i + 2, column=1, value=branches[i] if i < len(branches) else None)
            ws.cell(row=i + 2, column=2, value=departments[i] if i < len(departments) else None)
            ws.cell(row=i + 2, column=3, value=job_titles[i] if i < len(job_titles) else None)

        ws.freeze_panes = 'A2'

    # ─────────────────────────────────────────
    def _create_lists_sheet(self, wb):
        ws = wb.create_sheet("القوائم")

        from companies.models import Branch, Department
        from employees.models import Employee, JobTitle

        header_fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center      = Alignment(horizontal="center")

        lists = {
            "A": ("operation_type",     ["new", "update"]),
            "B": ("gender",             ["male", "female"]),
            "C": ("marital_status",     ["single", "married", "divorced", "widowed"]),
            "D": ("religion",           ["muslim", "christian", "other"]),
            "E": ("language",           ["ar", "en"]),
            "F": ("branch_name",        []),
            "G": ("department_name",    []),
            "H": ("job_title_name",     []),
            "I": ("attendance_mode",    ["fixed_shift", "flexible_hours", "field_worker", "multi_site", "rotating"]),
            "J": ("contract_type",      ["permanent", "temporary", "training", "freelance", "part_time", "consultant"]),
            "K": ("currency",           ["EGP", "USD", "SAR", "AED", "KWD", "QAR"]),
            "L": ("salary_payment_method", ["cash", "bank", "instapay", "wallet"]),
            "M": ("wallet_provider",    ["vodafone_cash", "orange_money", "etisalat_cash", "we_pay", "fawry", "other"]),
            "N": ("direct_manager_name",[]),
            "O": ("status",             ["active", "inactive", "terminated", "resigned", "on_leave"]),
            "P": ("worker_type",        ["مكتبي", "ميداني حر", "ميداني محدد"]),
        }

        # Populate from DB
        try:
            lists["F"] = ("branch_name",     list(Branch.objects.values_list("name_ar", flat=True)))
        except Exception:
            pass
        try:
            lists["G"] = ("department_name", list(Department.objects.values_list("name_ar", flat=True)))
        except Exception:
            pass
        try:
            lists["H"] = ("job_title_name",  list(JobTitle.objects.values_list("name_ar", flat=True)))
        except Exception:
            pass
        try:
            # المدير المباشر = أسماء الموظفين الكاملة
            mgr_names = []
            for emp in Employee.objects.select_related("user").all():
                name = f"{emp.first_name_ar} {emp.last_name_ar}".strip()
                if name:
                    mgr_names.append(name)
            lists["N"] = ("direct_manager_name", mgr_names)
        except Exception:
            pass

        for col_letter, (header, values) in lists.items():
            col_idx = ord(col_letter) - ord("A") + 1
            ws.column_dimensions[col_letter].width = 28

            header_cell = ws.cell(row=1, column=col_idx, value=header)
            header_cell.fill      = header_fill
            header_cell.font      = header_font
            header_cell.alignment = center

            for row_idx, value in enumerate(values, start=2):
                ws.cell(row=row_idx, column=col_idx, value=value)

        ws.sheet_state = "hidden"
