"""
Management Command: import_employees_bulk - Version 2
خريطة الأعمدة متطابقة مع generate_employee_template v2
"""
import os
import random
import string
from decimal import Decimal, InvalidOperation
from django.core.management.base import BaseCommand
from django.utils import timezone
from django.db import transaction
from openpyxl import load_workbook

from accounts.models import User
from employees.models import Employee, JobTitle
from companies.models import Branch, Department
from leaves.models import LeaveType, LeaveBalance
from leaves.signals import _get_entitlement_days
from employees.account_utils import send_mail
from django.conf import settings



# Mapping تصنيف الموظف من عربي لإنجليزي
WORKER_TYPE_MAP = {
    "مكتبي": "office",
    "ميداني حر": "field_free",
    "ميداني معين": "field_assigned",
    "ميداني محدد": "field_assigned",
    "office": "office",
    "field_free": "field_free",
    "field_assigned": "field_assigned",
}

# ═══════════════════════════════════════════════════════
# خريطة الأعمدة الثابتة — index يبدأ من 0
# ═══════════════════════════════════════════════════════
C = {
    "operation_type": 0,
    "full_name_ar": 1,
    "full_name_en": 2,
    "national_id": 3,
    "phone": 4,
    "birth_date": 5,
    "hire_date": 6,
    "branch_name": 7,
    "department_name": 8,
    "job_title_name": 9,
    "worker_type": 10,
    "basic_salary": 11,
    "direct_manager_name": 12,
}

OP_TYPE_MAP = {
    "جديد": "new",
    "تحديث": "update",
    "new": "new",
    "update": "update",
}



def _parse_flexible_date(value):
    """يقرأ التاريخ بأي فورمات ممكنة"""
    if value is None or value == '':
        return None

    # لو datetime object من openpyxl
    from datetime import datetime, date
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    # لو رقم (Excel serial date)
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(value).date()
        except Exception:
            pass

    # لو string - نجرب فورمات مختلفة
    s = str(value).strip()
    if not s:
        return None

    formats = [
        '%Y-%m-%d',
        '%d/%m/%Y',
        '%d-%m-%Y',
        '%m/%d/%Y',
        '%Y/%m/%d',
        '%d.%m.%Y',
        '%Y.%m.%d',
        '%d %m %Y',
        '%Y-%m-%d %H:%M:%S',
        '%d/%m/%Y %H:%M:%S',
    ]

    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    return None


def _get_val(row, key):
    """يجلب القيمة الخام بدون تحويل لـ string"""
    idx = C.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _split_full_name(full_name):
    """يقسم الاسم الكامل لـ (first, middle, last)"""
    if not full_name:
        return ("", "", "")
    parts = str(full_name).strip().split()
    if len(parts) == 0:
        return ("", "", "")
    if len(parts) == 1:
        return (parts[0], "", parts[0])
    if len(parts) == 2:
        return (parts[0], "", parts[1])
    return (parts[0], " ".join(parts[1:-1]), parts[-1])



LEAVE_COLS = {
    "annual":    ("annual_entitled",    "annual_used_before_system",    "annual_carry_forward"),
    "sick":      ("sick_entitled",      "sick_used_before_system",      "sick_carry_forward"),
    "emergency": ("emergency_entitled", "emergency_used_before_system", "emergency_carry_forward"),
    "maternity": ("maternity_entitled", "maternity_used_before_system", "maternity_carry_forward"),
    "paternity": ("paternity_entitled", "paternity_used_before_system", "paternity_carry_forward"),
    "unpaid":    ("unpaid_entitled",    "unpaid_used_before_system",    "unpaid_carry_forward"),
}


def _str(row, key):
    """يرجع string فارغة لو الحقل مش موجود (بدل ما يرمي error)"""
    idx = C.get(key)
    if idx is None:
        return ""
    if idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    return str(v).strip()


def _dec(row, key):
    val = row[C[key]]
    if val is None or str(val).strip() == "":
        return Decimal("0")
    try:
        return Decimal(str(val).strip())
    except InvalidOperation:
        return Decimal("0")


def _int(row, key, default=None):
    """يقرأ رقم صحيح - يرجع default لو الحقل مش موجود"""
    idx = C.get(key)
    if idx is None or idx >= len(row):
        return default
    v = row[idx]
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def _date(row, key, default=None):
    """يقرأ تاريخ - يرجع default لو الحقل مش موجود"""
    idx = C.get(key)
    if idx is None or idx >= len(row):
        return default
    v = row[idx]
    if v is None or v == "":
        return default
    return _parse_flexible_date(v) or default


def _company_employee_limit(company):
    """
    يرجع:
    - subscription
    - max_employees
    - current_active_employees
    لو مفيش باقة/حد، يرجع None
    """
    try:
        from subscriptions.models import CompanySubscription
        sub = (
            CompanySubscription.objects
            .filter(company=company, status__in=("trial", "active"))
            .select_related("plan")
            .first()
        )
        if not sub or not sub.plan:
            return None, None, None

        max_emp = sub.custom_max_employees or sub.plan.max_employees
        if max_emp is None:
            return sub, None, None

        current = Employee._base_manager.filter(company=company, status="active").count()
        return sub, max_emp, current
    except Exception:
        return None, None, None


class Command(BaseCommand):
    help = "استيراد الموظفين من ملف Excel - v2"

    def add_arguments(self, parser):
        parser.add_argument("--file",        type=str, required=True)
        parser.add_argument("--send-emails", action="store_true")
        parser.add_argument("--company-id",  type=int, default=None)

    def generate_random_password(self, length=8):
        chars = string.ascii_letters + string.digits
        return "".join(random.choice(chars) for _ in range(length))

    def _make_activation_link(self, user):
        from django.contrib.auth.tokens import default_token_generator
        from django.utils.http import urlsafe_base64_encode
        from django.utils.encoding import force_bytes
        uid   = urlsafe_base64_encode(force_bytes(user.pk))
        token = default_token_generator.make_token(user)
        site  = getattr(settings, "SITE_URL", "https://motion.jssolutions-eg.com")
        return f"{site}/password-reset-confirm/{uid}/{token}/"

    def handle(self, *args, **options):
        file_path  = options["file"]
        send_emails = options["send_emails"]

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"الملف غير موجود: {file_path}"))
            return

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb["الموظفين"]
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"خطأ في قراءة ملف الإكسيل: {e}"))
            return

        from companies.models import Company
        if options["company_id"]:
            company = Company.objects.filter(id=options["company_id"]).first()
        else:
            company = Company.objects.first()

        if not company:
            self.stdout.write(self.style.ERROR("لا توجد شركة مسجلة في النظام!"))
            return

        leave_types_map = {
            lt.category: lt
            for lt in LeaveType.objects.filter(company=company, is_active=True)
        }

        # تخطي أول 3 صفوف (section header + labels + keys)
        rows = list(ws.iter_rows(min_row=4, values_only=True))

        # Trial employee limit pre-check
        sub, max_emp, current_emp_count = _company_employee_limit(company)
        if max_emp is not None:
            new_rows_count = sum(
                1
                for row in rows
                if row and str(row[C["operation_type"]] or "").strip().lower() == "new"
            )
            remaining_slots = max(max_emp - current_emp_count, 0)

            if new_rows_count > remaining_slots:
                self.stdout.write(self.style.ERROR(
                    f"عدد الموظفين الجدد في الملف ({new_rows_count}) أكبر من المتاح في باقة الشركة "
                    f"({remaining_slots} متاح من أصل {max_emp}). "
                    f"يرجى الترقية أو تقليل عدد الصفوف الجديدة."
                ))
                return

        success_count = 0
        update_count  = 0
        error_count   = 0
        errors        = []

        for idx, row in enumerate(rows, start=4):
            if not row or not row[C["operation_type"]]:
                continue

            raw_op = _str(row, "operation_type").lower()
            op_type = OP_TYPE_MAP.get(raw_op, raw_op)
            emp_code = (_str(row, "employee_code") if "employee_code" in C else "")
            nat_id   = _str(row, "national_id")

            row_errors = self._validate_row(row, op_type, idx)
            if row_errors:
                for e in row_errors:
                    self.stdout.write(self.style.WARNING(e))
                    errors.append(e)
                error_count += 1
                continue

            try:
                with transaction.atomic():
                    if op_type == "new":
                        result = self._create_employee(
                            row, company, leave_types_map,
                            send_emails, idx
                        )
                        if result:
                            success_count += 1
                        else:
                            error_count += 1

                    elif op_type == "update":
                        result = self._update_employee(row, company, emp_code, nat_id, idx)
                        if result:
                            update_count += 1
                        else:
                            error_count += 1
                    else:
                        msg = f"صف {idx}: نوع العملية غير معروف: {op_type}"
                        self.stdout.write(self.style.WARNING(msg))
                        errors.append(msg)
                        error_count += 1

            except Exception as e:
                msg = f"صف {idx}: فشل الاستيراد - {e}"
                self.stdout.write(self.style.ERROR(msg))
                errors.append(msg)
                error_count += 1

        self.stdout.write(self.style.SUCCESS(
            f"تم الانتهاء! جديد: {success_count} | تحديث: {update_count} | أخطاء: {error_count}"
        ))
        if errors:
            self.stdout.write(self.style.WARNING("\nتفاصيل الأخطاء:"))
            for e in errors:
                self.stdout.write(f"  - {e}")

    # ─────────────────────────────────────────
    def _validate_row(self, row, op_type, idx):
        errors = []
        prefix = f"صف {idx}"

        if op_type == "new":
            # الحقول الجديدة المبسطة
            required_new = [
                ("full_name_ar", "الاسم الكامل بالعربي"),
                ("full_name_en", "الاسم الكامل بالإنجليزي"),
                ("national_id", "الرقم القومي"),
                ("phone", "الموبايل"),
                ("birth_date", "تاريخ الميلاد"),
                ("hire_date", "تاريخ التعيين"),
                ("branch_name", "الفرع"),
                ("department_name", "القسم"),
                ("job_title_name", "المسمى الوظيفي"),
                ("worker_type", "تصنيف الموظف"),
            ]
            for field, label in required_new:
                if not _str(row, field):
                    errors.append(f"{prefix}: الحقل [{label}] إجباري للموظف الجديد")

            worker_type = WORKER_TYPE_MAP.get(_str(row, "worker_type"), "")
            if worker_type and worker_type not in ("office", "field_free", "field_assigned"):
                errors.append(f"{prefix}: قيمة تصنيف الموظف غير صحيحة — المسموح: مكتبي / ميداني حر / ميداني محدد")

            return errors  # نتخطى باقي الـ validations القديمة

            worker_type = WORKER_TYPE_MAP.get(_str(row, "worker_type"), "")
            if worker_type and worker_type not in ("office", "field_free", "field_assigned"):
                errors.append(f"{prefix}: قيمة [worker_type] غير صحيحة — المسموح: مكتبي / ميداني حر / ميداني محدد")

            att_mode = _str(row, "attendance_mode")
            if att_mode and att_mode not in ("fixed_shift", "flexible_hours", "field_worker", "multi_site", "rotating"):
                errors.append(f"{prefix}: قيمة [attendance_mode] غير صحيحة")

            # التأمين
            has_ins = _str(row, "has_insurance")
            ins_num = _str(row, "insurance_number")
            if has_ins == "نعم" and not ins_num:
                errors.append(f"{prefix}: رقم التأمين إجباري لأن [مؤمن عليه = نعم]")

            # العقد
            contract_type = _str(row, "contract_type")
            if contract_type in ("temporary", "training", "consultant"):
                if not _str(row, "contract_start_date"):
                    errors.append(f"{prefix}: بداية العقد إجبارية لنوع العقد [{contract_type}]")
                if not _str(row, "contract_end_date") and not _int(row, "contract_duration_months"):
                    errors.append(f"{prefix}: نهاية العقد أو مدة العقد إجبارية لنوع العقد [{contract_type}]")

            # طريقة القبض
            pay_method = _str(row, "salary_payment_method")
            if pay_method == "bank":
                if not _str(row, "bank_name"):
                    errors.append(f"{prefix}: اسم البنك إجباري لطريقة القبض [bank]")
                if not _str(row, "bank_account"):
                    errors.append(f"{prefix}: رقم الحساب إجباري لطريقة القبض [bank]")

            elif pay_method == "instapay":
                if not _str(row, "instapay_transfer_id"):
                    errors.append(f"{prefix}: رقم إنستا باي إجباري لطريقة القبض [instapay]")
            elif pay_method == "wallet":
                if not _str(row, "wallet_transfer_number"):
                    errors.append(f"{prefix}: رقم المحفظة إجباري لطريقة القبض [wallet]")
                if not _str(row, "wallet_provider"):
                    errors.append(f"{prefix}: مزود المحفظة إجباري لطريقة القبض [wallet]")

        elif op_type == "update":
            if not (_str(row, "employee_code") if "employee_code" in C else "") and not _str(row, "national_id"):
                errors.append(f"{prefix}: كود الموظف أو الرقم القومي مطلوب للتحديث")

            worker_type = WORKER_TYPE_MAP.get(_str(row, "worker_type"), "")
            if worker_type and worker_type not in ("office", "field_free", "field_assigned"):
                errors.append(f"{prefix}: قيمة [worker_type] غير صحيحة — المسموح: مكتبي / ميداني حر / ميداني محدد")

            att_mode = _str(row, "attendance_mode")
            if att_mode and att_mode not in ("fixed_shift", "flexible_hours", "field_worker", "multi_site", "rotating"):
                errors.append(f"{prefix}: قيمة [attendance_mode] غير صحيحة")

        return errors

    # ─────────────────────────────────────────
    def _resolve_manager(self, company, manager_name, department=None, is_manager_job=False):
        """
        منطق ذكي لحل المدير المباشر:
        1. لو كُتب اسم مدير -> يبحث عنه ويطابقه.
        2. لو لم يُكتب اسم مدير:
           - لو موظف عادي -> يبحث عن مدير القسم، وإن لم يجد يربطه بـ HR أو صاحب الشركة.
           - لو مدير -> يربطه تلقائياً بصاحب الشركة / المدير العام.
        """
        # 1. إذا حُدد اسم مدير في الشيت
        if manager_name:
            # مطابقة بالاسم الكامل أو الاسم الأول والأخير
            employees = Employee._base_manager.filter(company=company).select_related("department", "user")
            for e in employees:
                e_full = f"{e.first_name_ar} {e.middle_name_ar} {e.last_name_ar}".strip().replace("  ", " ")
                e_short = f"{e.first_name_ar} {e.last_name_ar}".strip()
                if manager_name in [e_full, e_short, e.employee_code]:
                    return e
            # محاولة بحث تقريبية
            first_word = manager_name.split()[0] if manager_name.split() else manager_name
            candidate = Employee._base_manager.filter(company=company, first_name_ar__icontains=first_word).first()
            if candidate:
                return candidate

        # 2. في حالة عدم تحديد مدير (تحديد تلقائي ذكي)
        # البحث عن صاحب الشركة أو HR Manager
        owner_emp = Employee._base_manager.filter(
            company=company,
            user__role__in=['company_admin', 'super_admin']
        ).first()

        hr_emp = Employee._base_manager.filter(
            company=company,
            user__role='hr_manager'
        ).first()

        # لو الموظف هو نفسه مدير
        if is_manager_job:
            return owner_emp or hr_emp

        # لو موظف عادي: نبحث أولاً عن مدير القسم
        if department:
            dept_mgr = Employee._base_manager.filter(
                company=company,
                department=department,
                job_title__is_manager=True
            ).first()
            if dept_mgr:
                return dept_mgr

        # إن لم يوجد مدير قسم -> HR أو صاحب الشركة
        return hr_emp or owner_emp
        matches = Employee._base_manager.filter(
            company=company,
            first_name_ar__icontains=manager_name.split()[0] if manager_name.split() else manager_name,
        )
        full_matches = [
            e for e in matches
            if f"{e.first_name_ar} {e.last_name_ar}".strip() == manager_name
        ]
        if len(full_matches) == 1:
            return full_matches[0]
        elif len(full_matches) > 1:
            raise ValueError(f"اسم المدير [{manager_name}] متكرر — يرجى التوضيح بكود الموظف")
        return None

    # ─────────────────────────────────────────
    def _create_employee(self, row, company, leave_types_map, send_emails, idx):
        # قراءة الأسماء الجديدة (full_name_ar / full_name_en) + fallback للقديم
        full_ar = _str(row, "full_name_ar")
        full_en = _str(row, "full_name_en")
        print(f"[DEBUG] Row {idx}: full_ar='{full_ar}' full_en='{full_en}'")

        if full_ar:
            fname_ar, mname_ar, lname_ar = _split_full_name(full_ar)
        else:
            fname_ar = _str(row, "first_name_ar")
            mname_ar = _str(row, "middle_name_ar")
            lname_ar = _str(row, "last_name_ar")

        if full_en:
            fname_en, _mid_en, lname_en = _split_full_name(full_en)
        else:
            fname_en = _str(row, "first_name_en")
            lname_en = _str(row, "last_name_en")


        nat_id    = _str(row, "national_id")
        phone     = _str(row, "phone")
        email     = _str(row, "email") or None
        emp_code  = (_str(row, "employee_code") if "employee_code" in C else "")
        temp_pass = _str(row, "temporary_password") or self.generate_random_password()

        hire_date = _date(row, "hire_date")

        branch_name = _str(row, "branch_name")
        dept_name   = _str(row, "department_name")
        job_name    = _str(row, "job_title_name")

        created_defs = []

        branch = Branch._base_manager.filter(company=company, name_ar=branch_name).first() if branch_name else None
        if not branch and branch_name:
            branch = Branch._base_manager.create(company=company, name_ar=branch_name)
            created_defs.append(f"صف {idx}: تم إنشاء فرع جديد [{branch_name}]")

        dept = Department._base_manager.filter(company=company, name_ar=dept_name).first() if dept_name else None
        if not dept and dept_name:
            dept = Department._base_manager.create(company=company, name_ar=dept_name)
            created_defs.append(f"صف {idx}: تم إنشاء قسم جديد [{dept_name}]")

        job = JobTitle._base_manager.filter(company=company, name_ar=job_name).first() if job_name else None
        if not job and job_name:
            job = JobTitle._base_manager.create(company=company, name_ar=job_name)
            created_defs.append(f"صف {idx}: تم إنشاء مسمى وظيفي جديد [{job_name}]")

        manager = self._resolve_manager(company, _str(row, "direct_manager_name"), department=dept, is_manager_job=getattr(job, "is_manager", False))

        contract_type = _str(row, "contract_type") or "permanent"
        contract_start = _date(row, "contract_start_date") or hire_date
        contract_end   = _date(row, "contract_end_date")
        duration_months = _int(row, "contract_duration_months")

        # حساب نهاية العقد تلقائي لو مش موجودة
        if not contract_end and duration_months and contract_start:
            from dateutil.relativedelta import relativedelta
            contract_end = contract_start + relativedelta(months=duration_months)

        has_insurance = _str(row, "has_insurance") == "نعم"
        gender_val = _str(row, "gender")

        att_mode = _str(row, "attendance_mode") or "fixed_shift"
        worker_type_val = WORKER_TYPE_MAP.get(_str(row, "worker_type"), "")
        if worker_type_val not in ("office", "field_free", "field_assigned"):
            raise ValueError("حقل [تصنيف الموظف] إجباري — اختر واحدة من: مكتبي / ميداني حر / ميداني محدد")
        status   = "active"  # دايمًا نشط - مش بنقرأ من الإكسيل

        # Trial employee limit runtime guard
        _, max_emp, current_emp_count = _company_employee_limit(company)
        if max_emp is not None and current_emp_count >= max_emp:
            raise ValueError(
                f"تم الوصول للحد الأقصى لعدد الموظفين في الباقة الحالية ({max_emp})"
            )

        username = f"emp{nat_id[-6:]}{random.randint(10, 99)}"
        user = User.objects.create_user(
            username=username,
            password=temp_pass,
            email=email or "",
            first_name=fname_ar,
            last_name=lname_ar,
            company=company,
            role="employee",
        )

        emp = Employee.objects.create(
            company=company,
            user=user,
            employee_code=emp_code or f"EMP{user.id}",
            first_name_ar=fname_ar,
            middle_name_ar=mname_ar,
            last_name_ar=lname_ar,
            first_name_en=fname_en,
            last_name_en=lname_en,
            national_id=nat_id,
            birth_date=_date(row, "birth_date"),
            gender="female" if gender_val in ("female", "أنثى") else "male",
            marital_status=_str(row, "marital_status") or "single",
            religion=_str(row, "religion") or "muslim",
            nationality=_str(row, "nationality") or "مصري",
            language="en" if _str(row, "language") == "en" else "ar",
            phone=phone,
            phone2=_str(row, "phone2") or None,
            email=email,
            address=_str(row, "address") or None,
            city=_str(row, "city") or None,
            emergency_contact_name=_str(row, "emergency_contact_name") or None,
            emergency_contact_relation=_str(row, "emergency_contact_relation") or None,
            emergency_contact_phone=_str(row, "emergency_contact_phone") or None,
            branch=branch,
            department=dept,
            job_title=job,
            direct_manager=manager,
            hire_date=hire_date,
            attendance_mode=att_mode,
            status=status,
            worker_type=worker_type_val,
            contract_type=contract_type,
            contract_start_date=contract_start,
            contract_end_date=contract_end,
            contract_duration_months=duration_months or None,
            probation_months=_int(row, "probation_months", 3),
            has_insurance=has_insurance,
            insurance_number=_str(row, "insurance_number") or None,
            basic_salary=_dec(row, "basic_salary"),
            currency=_str(row, "currency") or "EGP",
            salary_payment_method=_str(row, "salary_payment_method") or "cash",
            bank_name=_str(row, "bank_name") or None,
            bank_account=_str(row, "bank_account") or None,
            iban=_str(row, "iban") or None,
            instapay_phone=_str(row, "instapay_transfer_id") or None,
            wallet_phone=_str(row, "wallet_transfer_number") or None,
            wallet_provider=_str(row, "wallet_provider") or None,
        )

        # أرصدة الإجازات
        year = timezone.now().year
        for cat, (ent_key, used_key, carry_key) in LEAVE_COLS.items():
            lt = leave_types_map.get(cat)
            if not lt:
                continue
            entitled = _dec(row, ent_key)
            used     = _dec(row, used_key)
            carry    = _dec(row, carry_key)
            if entitled == 0:
                entitled = Decimal(str(_get_entitlement_days(company, emp, lt, year)))
            LeaveBalance._base_manager.create(
                company=company,
                employee=emp,
                leave_type=lt,
                year=year,
                total_days=entitled + carry,
                used_days=used,
                pending_days=0,
            )

        # ═══════════════════════════════════════════════════
        # ربط الموظف بالشيفت الافتراضي للشركة (Auto-link)
        # ═══════════════════════════════════════════════════
        try:
            from attendance.models import Shift, EmployeeShift

            default_shift = Shift._base_manager.filter(
                company=company,
                is_default=True,
                is_active=True,
            ).first()

            if default_shift:
                already_linked = EmployeeShift._base_manager.filter(
                    company=company,
                    employee=emp,
                    shift=default_shift,
                    is_active=True,
                ).exists()

                if not already_linked:
                    EmployeeShift._base_manager.create(
                        company=company,
                        employee=emp,
                        shift=default_shift,
                        assignment_type="individual",
                        start_date=timezone.now().date(),
                        is_active=True,
                        priority=1,
                    )
                    created_defs.append(
                        f"صف {idx}: تم ربط الموظف بالشيفت الافتراضي [{default_shift.name}]"
                    )
        except Exception as e:
            self.stdout.write(self.style.WARNING(
                f"صف {idx}: تعذر ربط الشيفت الافتراضي: {e}"
            ))

        for msg in created_defs:
            self.stdout.write(self.style.SUCCESS(msg))

        self.stdout.write(self.style.SUCCESS(f"صف {idx}: تم إنشاء الموظف [{fname_ar} {lname_ar}] بنجاح"))

        # إرسال الإيميل
        if send_emails and email:
            try:
                act_link = self._make_activation_link(user)
                msg = (
                    f"مرحباً {fname_ar}،\n"
                    f"تم إنشاء حسابك في MotionHR.\n"
                    f"اسم المستخدم: {username}\n"
                    f"كلمة المرور المؤقتة: {temp_pass}\n"
                    f"رابط التفعيل (صالح 48 ساعة): {act_link}"
                )
                send_mail(
                    "مرحباً بك في MotionHR",
                    msg,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=True,
                )
            except Exception:
                pass

        return True

    # ─────────────────────────────────────────
    def _update_employee(self, row, company, emp_code, nat_id, idx):
        emp = None
        if emp_code:
            emp = Employee._base_manager.filter(employee_code=emp_code, company=company).first()
        if not emp and nat_id:
            emp = Employee._base_manager.filter(national_id=nat_id, company=company).first()

        if not emp:
            msg = f"صف {idx}: لم يتم العثور على الموظف (كود: {emp_code} | رقم قومي: {nat_id})"
            self.stdout.write(self.style.WARNING(msg))
            return False

        # تحديث الحقول اللي موجودة في الصف
        salary = _dec(row, "basic_salary")
        if salary:
            emp.basic_salary = salary

        contract_type = _str(row, "contract_type")
        if contract_type:
            emp.contract_type = contract_type

        # الحالة الوظيفية لا تتحدث من الإكسيل
        # الموظف يفضل نشط إلا بتغيير إداري مباشر

        att_mode = _str(row, "attendance_mode")
        if att_mode:
            emp.attendance_mode = att_mode

        worker_type_val = WORKER_TYPE_MAP.get(_str(row, "worker_type"), "")
        if worker_type_val in ("office", "field_free", "field_assigned"):
            emp.worker_type = worker_type_val

        contract_start = _date(row, "contract_start_date")
        if contract_start:
            emp.contract_start_date = contract_start

        contract_end = _date(row, "contract_end_date")
        if contract_end:
            emp.contract_end_date = contract_end

        duration = _int(row, "contract_duration_months")
        if duration:
            emp.contract_duration_months = duration

        probation = _int(row, "probation_months")
        if probation:
            emp.probation_months = probation

        pay_method = _str(row, "salary_payment_method")
        if pay_method:
            emp.salary_payment_method = pay_method

        manager = self._resolve_manager(company, _str(row, "direct_manager_name"), department=dept, is_manager_job=getattr(job, "is_manager", False))
        if manager:
            emp.direct_manager = manager

        emp.save()
        self.stdout.write(self.style.SUCCESS(f"صف {idx}: تم تحديث الموظف [{emp.first_name_ar} {emp.last_name_ar}]"))
        return True
