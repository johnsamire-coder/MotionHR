"""
reports/utils.py — Patch 49a Fix3
الـ field names الصح:
  check_in_time / check_out_time
  company.name_ar
  department.name_ar
"""

from datetime import date, timedelta, datetime
from django.http import HttpResponse
from django.utils import timezone

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ════════════════════════════════════
# Date Range
# ════════════════════════════════════

def get_date_range(period, custom_start=None, custom_end=None):
    today = date.today()
    if period == "week":
        start = today - timedelta(days=today.weekday())
        end   = start + timedelta(days=6)
    elif period == "month":
        start = today.replace(day=1)
        next_m = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
        end   = next_m - timedelta(days=1)
    elif period == "quarter":
        q     = (today.month - 1) // 3
        start = date(today.year, q * 3 + 1, 1)
        end   = (date(today.year, (q+1)*3+1, 1) - timedelta(days=1)) if q < 3 else date(today.year, 12, 31)
    elif period == "year":
        start = today.replace(month=1, day=1)
        end   = today.replace(month=12, day=31)
    elif period == "custom" and custom_start and custom_end:
        start, end = custom_start, custom_end
    else:
        start = today.replace(day=1)
        end   = today
    return start, end


# ════════════════════════════════════
# Attendance Summary
# ════════════════════════════════════

def get_attendance_summary(company, start_date, end_date):
    from attendance.models import Attendance
    qs      = Attendance.objects.filter(company=company, date__range=(start_date, end_date))
    total   = qs.count()
    present = qs.filter(status='present').count()
    late    = qs.filter(status='late').count()
    absent  = qs.filter(status='absent').count()
    on_leave= qs.filter(status='on_leave').count()
    return {
        'total':    total,
        'present':  present,
        'late':     late,
        'absent':   absent,
        'on_leave': on_leave,
        'rate':     round(present / total * 100, 1) if total else 0,
    }


# ════════════════════════════════════
# Employee Attendance Details
# ════════════════════════════════════

def get_employee_attendance_details(company, start_date, end_date):
    from attendance.models import Attendance
    from employees.models import Employee
    from django.db.models import Sum

    employees = Employee.objects.filter(company=company, status='active')
    result = []

    for emp in employees:
        qs      = Attendance.objects.filter(employee=emp, date__range=(start_date, end_date))
        total   = qs.count()
        present = qs.filter(status='present').count()
        late    = qs.filter(status='late').count()
        absent  = qs.filter(status='absent').count()
        on_leave= qs.filter(status='on_leave').count()
        late_mins = qs.aggregate(Sum('late_minutes'))['late_minutes__sum'] or 0

        # ساعات العمل من work_hours field
        from django.db.models import Sum as S
        work_hours_sum = qs.aggregate(S('work_hours'))['work_hours__sum']
        if work_hours_sum is None:
            # احسب يدوي من check_in_time و check_out_time
            work_seconds = 0
            for a in qs:
                if a.check_in_time and a.check_out_time:
                    ci = datetime.combine(date.today(), a.check_in_time)
                    co = datetime.combine(date.today(), a.check_out_time)
                    if co > ci:
                        work_seconds += (co - ci).seconds
            work_hours = round(work_seconds / 3600, 1)
        else:
            work_hours = round(float(work_hours_sum), 1)

        result.append({
            'employee':   emp,
            'total':      total,
            'present':    present,
            'late':       late,
            'absent':     absent,
            'on_leave':   on_leave,
            'rate':       round(present / total * 100, 1) if total else 0,
            'late_mins':  late_mins,
            'work_hours': work_hours,
        })
    return result


# ════════════════════════════════════
# Field Tracking Summary
# ════════════════════════════════════

def get_field_tracking_summary(company, start_date, end_date):
    from attendance.models import LocationLog
    from employees.models import Employee

    employees = Employee.objects.filter(company=company, status='active', is_field_worker=True)
    result = []
    for emp in employees:
        logs = LocationLog.objects.filter(
            employee=emp,
            timestamp__date__range=(start_date, end_date)
        )
        result.append({
            'employee':  emp,
            'log_count': logs.count(),
            'last_log':  logs.order_by('-timestamp').first(),
        })
    return result


# ════════════════════════════════════
# Excel Export — openpyxl احترافي
# ════════════════════════════════════

def export_to_excel(headers, rows, filename="report", company=None, report_title="تقرير"):
    if not OPENPYXL_AVAILABLE:
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        import csv
        writer = csv.writer(response)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_title[:31]
    ws.sheet_view.rightToLeft = True

    PRIMARY   = "06B6D4"
    HEADER_BG = "0E7490"
    LIGHT_ROW = "F0FDFF"
    WHITE     = "FFFFFF"
    DARK      = "1E293B"
    GRAY      = "94A3B8"

    thin_border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6'),
    )

    num_cols = max(len(headers), 1)
    last_col = get_column_letter(num_cols)
    current_row = 1

    # ── اسم الشركة ──
    company_name = ""
    if company:
        company_name = getattr(company, 'name_ar', None) or getattr(company, 'name_en', None) or "MotionHR"
    else:
        company_name = "MotionHR"

    ws.merge_cells(f'A{current_row}:{last_col}{current_row}')
    cell = ws.cell(row=current_row, column=1, value=company_name)
    cell.font      = Font(name='Cairo', size=16, bold=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 35
    current_row += 1

    # ── عنوان التقرير ──
    ws.merge_cells(f'A{current_row}:{last_col}{current_row}')
    cell = ws.cell(row=current_row, column=1, value=report_title)
    cell.font      = Font(name='Cairo', size=13, bold=True, color=DARK)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # ── تاريخ التصدير ──
    ws.merge_cells(f'A{current_row}:{last_col}{current_row}')
    export_date = timezone.now().strftime("%Y/%m/%d %H:%M")
    cell = ws.cell(row=current_row, column=1, value=f"تاريخ التصدير: {export_date}")
    cell.font      = Font(name='Cairo', size=10, color=GRAY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 20
    current_row += 2

    # ── Header ──
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.font      = Font(name='Cairo', size=11, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_border
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # ── Data ──
    for row_idx, row in enumerate(rows):
        bg = LIGHT_ROW if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name='Cairo', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border    = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # ── Footer ──
    current_row += 1
    ws.merge_cells(f'A{current_row}:{last_col}{current_row}')
    cell = ws.cell(row=current_row, column=1,
                   value="MotionHR — HR in Motion | JS Solutions")
    cell.font      = Font(name='Cairo', size=9, color=GRAY, italic=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # ── عرض الأعمدة ──
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for c in row:
                try:
                    if c.value:
                        max_len = max(max_len, len(str(c.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response
