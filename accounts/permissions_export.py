"""
MotionHR — Unified Permissions Export
Design: Navy (#1A1B4B) + Cyan (#06B6D4) + Company Logo + Bilingual (AR + EN)
"""
from pathlib import Path
from io import BytesIO
from datetime import datetime

import arabic_reshaper
from bidi.algorithm import get_display
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage,
    PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.lib.units import mm

from .permissions_models import (
    CustomRole, UserRole, UserPermissionOverride,
    PERMISSION_CHOICES, SCOPE_CHOICES
)

# ══════════════════════════════════════
# COLORS — MotionHR Brand
# ══════════════════════════════════════
NAVY = colors.HexColor("#1A1B4B")
CYAN = colors.HexColor("#06B6D4")
LIGHT_GRAY = colors.HexColor("#F3F4F6")
DARK_GRAY = colors.HexColor("#6B7280")
BORDER_GRAY = colors.HexColor("#E5E7EB")

NAVY_HEX = "1A1B4B"
CYAN_HEX = "06B6D4"
LIGHT_GRAY_HEX = "F3F4F6"
DARK_GRAY_HEX = "6B7280"
BORDER_GRAY_HEX = "E5E7EB"

# ══════════════════════════════════════
# Helper: ترجمة الصلاحية والنطاق
# ══════════════════════════════════════
PERM_MAP = dict(PERMISSION_CHOICES)
SCOPE_MAP = dict(SCOPE_CHOICES)

# EN labels (fallback to code)
PERM_EN = {
    'employees.view': 'View Employees',
    'employees.add': 'Add Employee',
    'employees.edit': 'Edit Employee',
    'employees.delete': 'Delete Employee',
    'employees.transfer': 'Transfer Employee',
    'attendance.view': 'View Attendance',
    'attendance.edit': 'Edit Attendance',
    'leaves.view': 'View Leaves',
    'leaves.approve': 'Approve Leaves',
    'requests.view': 'View Requests',
    'requests.approve': 'Approve Requests',
    'payroll.view': 'View Payroll',
    'payroll.edit': 'Edit Payroll',
    'reports.view': 'View Reports',
    'reports.export': 'Export Reports',
    'missions.view': 'View Missions',
    'missions.manage': 'Manage Missions',
    'company.view': 'View Company Settings',
    'company.edit': 'Edit Company Settings',
    'departments.view': 'View Departments',
    'departments.add': 'Add Department',
    'departments.edit': 'Edit Department',
    'departments.delete': 'Delete Department',
    'departments.transfer_employees': 'Transfer Between Departments',
    'offboarding.execute': 'Offboard Employee/Manager',
    'shifts.view': 'View Shifts',
    'shifts.manage': 'Manage Shifts',
    'policies.view': 'View Policies',
    'policies.manage': 'Manage Policies',
    'holidays.view': 'View Holidays',
    'holidays.manage': 'Manage Holidays',
    'tracking.view': 'View Tracking',
    'tracking.manage': 'Manage Tracking',
    'payroll.view_own': 'View Own Payroll',
    'attendance.checkin': 'Attendance Check-in',
    'leaves.request': 'Request Leave',
    'requests.submit': 'Submit Request',
    'profile.view': 'View Profile',
    'profile.edit_basic': 'Edit Basic Profile',
    'missions.view_own': 'View Own Missions',
    'roles.manage': 'Manage Roles & Permissions',
}

SCOPE_EN = {
    'company': 'Company',
    'branch': 'Branch',
    'department': 'Department',
    'self': 'Self',
    'all': 'All',
}

def perm_label(code):
    return PERM_MAP.get(code, code)

def perm_label_en(code):
    return PERM_EN.get(code, code)

def scope_label(code):
    return SCOPE_MAP.get(code, code)

def scope_label_en(code):
    return SCOPE_EN.get(code, code)


# ══════════════════════════════════════
# PDF Arabic Helpers
# ══════════════════════════════════════
PDF_FONT_NAME = "NotoNaskhArabic"
PDF_FONT_PATH = Path("/usr/share/fonts/opentype/fonts-hosny-amiri/Amiri-Regular.ttf")
PDF_FONT_BOLD_PATH = Path("/usr/share/fonts/opentype/fonts-hosny-amiri/Amiri-Bold.ttf")
PDF_FONT_BOLD_NAME = "NotoNaskhArabicBold"


def _ensure_pdf_font():
    if PDF_FONT_NAME not in pdfmetrics.getRegisteredFontNames():
        if PDF_FONT_PATH.exists():
            pdfmetrics.registerFont(TTFont(PDF_FONT_NAME, str(PDF_FONT_PATH)))
    if PDF_FONT_BOLD_NAME not in pdfmetrics.getRegisteredFontNames():
        bold_path = PDF_FONT_BOLD_PATH if PDF_FONT_BOLD_PATH.exists() else PDF_FONT_PATH
        pdfmetrics.registerFont(TTFont(PDF_FONT_BOLD_NAME, str(bold_path)))


def _has_arabic(text: str) -> bool:
    return any('\u0600' <= ch <= '\u06FF' for ch in str(text))


def ar_text(value) -> str:
    """Reshape Arabic text for PDF."""
    text = str(value or '')
    if not text:
        return ''
    if _has_arabic(text):
        return get_display(arabic_reshaper.reshape(text))
    return text


def bilingual(ar_val, en_val=None) -> str:
    """Return 'AR — EN' format (or just AR if EN missing)."""
    ar = str(ar_val or '').strip()
    en = str(en_val or '').strip()
    if ar and en and ar != en:
        return f"{ar} / {en}"
    return ar or en


# ══════════════════════════════════════
# LOGO Helper
# ══════════════════════════════════════
def _get_logo_path(company):
    """Get company logo file path or None."""
    try:
        if company and hasattr(company, 'logo') and company.logo:
            path = company.logo.path
            if Path(path).exists():
                return path
    except Exception:
        pass
    return None


# ══════════════════════════════════════
# PDF Styles
# ══════════════════════════════════════
def _pdf_styles():
    _ensure_pdf_font()
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "title", parent=styles["Title"], fontName=PDF_FONT_BOLD_NAME,
            alignment=TA_CENTER, fontSize=18, leading=24, textColor=NAVY,
        ),
        "subtitle": ParagraphStyle(
            "subtitle", parent=styles["Normal"], fontName=PDF_FONT_NAME,
            alignment=TA_CENTER, fontSize=11, leading=14, textColor=DARK_GRAY,
        ),
        "heading": ParagraphStyle(
            "heading", parent=styles["Heading2"], fontName=PDF_FONT_BOLD_NAME,
            alignment=TA_RIGHT, fontSize=13, leading=18, textColor=NAVY,
            spaceBefore=8, spaceAfter=4,
        ),
        "body": ParagraphStyle(
            "body", parent=styles["Normal"], fontName=PDF_FONT_NAME,
            alignment=TA_RIGHT, fontSize=10, leading=14,
        ),
        "company_name": ParagraphStyle(
            "company_name", parent=styles["Normal"], fontName=PDF_FONT_BOLD_NAME,
            alignment=TA_RIGHT, fontSize=16, leading=20, textColor=colors.white,
        ),
        "company_sub": ParagraphStyle(
            "company_sub", parent=styles["Normal"], fontName=PDF_FONT_NAME,
            alignment=TA_RIGHT, fontSize=9, leading=12, textColor=colors.HexColor("#CBD5E1"),
        ),
    }


# ══════════════════════════════════════
# PDF Header & Footer
# ══════════════════════════════════════
def _build_pdf_header(company, subtitle_ar="", subtitle_en=""):
    """Build the unified PDF header (logo + company name + subtitle)."""
    styles = _pdf_styles()
    logo_path = _get_logo_path(company)

    company_name_ar = (getattr(company, 'name_ar', '') or '').strip()
    company_name_en = (getattr(company, 'name_en', '') or '').strip()

    # Logo cell
    if logo_path:
        try:
            logo = RLImage(logo_path, width=25*mm, height=25*mm, kind='proportional')
        except Exception:
            logo = Paragraph("<b>HR</b>", styles["company_name"])
    else:
        logo = Paragraph("<b>MotionHR</b>", styles["company_name"])

    # Company info cell
    company_info = []
    if company_name_ar:
        company_info.append(Paragraph(ar_text(company_name_ar), styles["company_name"]))
    if company_name_en:
        company_info.append(Paragraph(
            f'<font name="{PDF_FONT_NAME}">{company_name_en}</font>',
            styles["company_sub"]
        ))
    company_info.append(Spacer(1, 3))
    company_info.append(Paragraph(
        f'<font name="{PDF_FONT_NAME}">MotionHR Workforce Platform</font>',
        styles["company_sub"]
    ))

    header_table = Table(
        [[company_info, logo]],
        colWidths=[130*mm, 40*mm],
        rowHeights=[30*mm],
    )
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), NAVY),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))

    elements = [header_table, Spacer(1, 10)]

    # Report subtitle
    if subtitle_ar or subtitle_en:
        elements.append(Paragraph(ar_text(subtitle_ar or subtitle_en), styles["title"]))
        if subtitle_ar and subtitle_en:
            elements.append(Paragraph(
                f'<font name="{PDF_FONT_NAME}">{subtitle_en}</font>',
                styles["subtitle"]
            ))
        elements.append(Spacer(1, 8))

    # Date
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    elements.append(Paragraph(
        f'<font name="{PDF_FONT_NAME}">{now_str}</font>',
        styles["subtitle"]
    ))
    elements.append(Spacer(1, 12))
    return elements


def _add_page_footer(canvas, doc):
    """Draw footer with page number + branding."""
    canvas.saveState()
    page_num = canvas.getPageNumber()
    page_width, page_height = A4

    # Line
    canvas.setStrokeColor(BORDER_GRAY)
    canvas.setLineWidth(0.5)
    canvas.line(15*mm, 15*mm, page_width - 15*mm, 15*mm)

    # Left: Powered by
    canvas.setFont(PDF_FONT_NAME, 8)
    canvas.setFillColor(DARK_GRAY)
    canvas.drawString(15*mm, 10*mm, "Powered by MotionHR")

    # Right: Page number
    canvas.drawRightString(page_width - 15*mm, 10*mm, f"Page {page_num}")

    canvas.restoreState()


# ══════════════════════════════════════
# EXCEL Helpers
# ══════════════════════════════════════
def _style_excel_header(cell, bg=NAVY_HEX, color="FFFFFF", size=11, bold=True):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, color=color, size=size, name="Cairo")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color=BORDER_GRAY_HEX)
    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def _style_excel_data(cell, is_even=False):
    cell.fill = PatternFill("solid", fgColor="FFFFFF" if is_even else "F9FAFB")
    cell.font = Font(size=10, name="Cairo")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color=BORDER_GRAY_HEX)
    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def _add_excel_header(sheet, company, title_ar, title_en, cols_count=4):
    """Add unified header to Excel sheet with logo + company name."""
    sheet.sheet_view.rightToLeft = True

    # Logo
    logo_path = _get_logo_path(company)
    start_row = 1
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 80
            img.height = 80
            sheet.add_image(img, "A1")
            for i in range(1, 5):
                sheet.row_dimensions[i].height = 20
            start_row = 5
        except Exception:
            pass

    # Company name (AR + EN)
    company_name_ar = (getattr(company, 'name_ar', '') or '').strip()
    company_name_en = (getattr(company, 'name_en', '') or '').strip()

    row = start_row
    if company_name_ar:
        cell = sheet.cell(row=row, column=1, value=company_name_ar)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols_count)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY_HEX)
        cell.font = Font(bold=True, size=16, color=NAVY_HEX, name="Cairo")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[row].height = 30
        row += 1

    if company_name_en:
        cell = sheet.cell(row=row, column=1, value=company_name_en)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols_count)
        cell.font = Font(size=11, color=DARK_GRAY_HEX, italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    # Title (AR)
    title_text = title_ar or title_en
    if title_text:
        cell = sheet.cell(row=row, column=1, value=title_text)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols_count)
        cell.fill = PatternFill("solid", fgColor=NAVY_HEX)
        cell.font = Font(bold=True, size=14, color="FFFFFF", name="Cairo")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[row].height = 28
        row += 1

    # Title (EN)
    if title_en and title_ar and title_en != title_ar:
        cell = sheet.cell(row=row, column=1, value=title_en)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols_count)
        cell.fill = PatternFill("solid", fgColor=CYAN_HEX)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    # Date
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    cell = sheet.cell(row=row, column=1, value=now_str)
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols_count)
    cell.font = Font(size=9, color=DARK_GRAY_HEX, italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2  # blank line
    return row


def _add_excel_footer(sheet, row, cols_count=4):
    """Add footer with branding."""
    cell = sheet.cell(row=row, column=1, value="Generated by MotionHR Workforce Platform")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols_count)
    cell.font = Font(size=9, color=DARK_GRAY_HEX, italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ══════════════════════════════════════
# ROLE EXPORT — EXCEL
# ══════════════════════════════════════
def export_role_excel(role: CustomRole) -> HttpResponse:
    company = role.company
    wb = Workbook()
    ws = wb.active
    ws.title = "Role Permissions"

    # Header
    row = _add_excel_header(
        ws, company,
        title_ar=f"صلاحيات الدور: {role.name}",
        title_en=f"Role Permissions: {role.name}",
        cols_count=3,
    )

    # Table headers
    headers = ["الصلاحية / Permission", "الكود / Code", "النطاق / Scope"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        _style_excel_header(cell)
    ws.row_dimensions[row].height = 26
    row += 1

    # Data
    perms = role.permissions.all().order_by('permission')
    for idx, p in enumerate(perms):
        is_even = idx % 2 == 0
        vals = [
            f"{perm_label(p.permission)} / {perm_label_en(p.permission)}",
            p.permission,
            f"{scope_label(p.scope)} / {scope_label_en(p.scope)}",
        ]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            _style_excel_data(cell, is_even)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25

    # Footer
    row += 1
    _add_excel_footer(ws, row, cols_count=3)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = f'attachment; filename="role_{role.id}_permissions.xlsx"'
    return resp


# ══════════════════════════════════════
# USER EXPORT — EXCEL
# ══════════════════════════════════════
def export_user_excel(user) -> HttpResponse:
    company = user.company
    wb = Workbook()
    ws = wb.active
    ws.title = "User Permissions"

    user_name = f"{user.first_name} {user.last_name}".strip() or user.username

    row = _add_excel_header(
        ws, company,
        title_ar=f"صلاحيات المستخدم: {user_name}",
        title_en=f"User Permissions: {user_name}",
        cols_count=4,
    )

    # User info
    info_cells = [
        ("اسم المستخدم / Username:", user.username),
        ("البريد / Email:", user.email or "—"),
    ]
    for label, val in info_cells:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(bold=True, size=10, name="Cairo")
        c1.alignment = Alignment(horizontal="right", vertical="center")
        c2 = ws.cell(row=row, column=2, value=str(val))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        c2.font = Font(size=10, name="Cairo")
        c2.alignment = Alignment(horizontal="right", vertical="center")
        row += 1
    row += 1

    # Roles section
    cell = ws.cell(row=row, column=1, value="الأدوار المعينة / Assigned Roles")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell.fill = PatternFill("solid", fgColor=CYAN_HEX)
    cell.font = Font(bold=True, size=11, color="FFFFFF", name="Cairo")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    user_roles = UserRole.objects.filter(user=user).select_related('role')
    if user_roles.exists():
        for idx, ur in enumerate(user_roles):
            c = ws.cell(row=row, column=1, value=ur.role.name)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            _style_excel_data(c, idx % 2 == 0)
            c.alignment = Alignment(horizontal="right", vertical="center")
            row += 1
    else:
        c = ws.cell(row=row, column=1, value="— لا توجد أدوار / No roles —")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        _style_excel_data(c, True)
        row += 1
    row += 1

    # Overrides section
    cell = ws.cell(row=row, column=1, value="الاستثناءات / Overrides")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell.fill = PatternFill("solid", fgColor=CYAN_HEX)
    cell.font = Font(bold=True, size=11, color="FFFFFF", name="Cairo")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    # Overrides headers
    headers = ["الصلاحية / Permission", "النطاق / Scope", "الحالة / Status", "ملاحظات / Note"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        _style_excel_header(cell)
    ws.row_dimensions[row].height = 26
    row += 1

    overrides = UserPermissionOverride.objects.filter(user=user).order_by('permission')
    if overrides.exists():
        for idx, o in enumerate(overrides):
            is_even = idx % 2 == 0
            status_ar = "مسموح ✓" if o.is_granted else "ممنوع ✗"
            status_en = "Granted" if o.is_granted else "Denied"
            vals = [
                f"{perm_label(o.permission)} / {perm_label_en(o.permission)}",
                f"{scope_label(o.scope)} / {scope_label_en(o.scope)}",
                f"{status_ar} / {status_en}",
                o.note or "—",
            ]
            for i, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=i, value=v)
                _style_excel_data(cell, is_even)
            row += 1
    else:
        c = ws.cell(row=row, column=1, value="— لا توجد استثناءات / No overrides —")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        _style_excel_data(c, True)
        row += 1

    # Column widths
    for col, width in [('A', 35), ('B', 20), ('C', 20), ('D', 25)]:
        ws.column_dimensions[col].width = width

    # Footer
    row += 2
    _add_excel_footer(ws, row, cols_count=4)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = f'attachment; filename="user_{user.id}_permissions.xlsx"'
    return resp


# ══════════════════════════════════════
# COMPANY EXPORT — EXCEL
# ══════════════════════════════════════
def export_company_excel(company) -> HttpResponse:
    wb = Workbook()

    # ─── Sheet 1: All Roles ─────────────────
    ws1 = wb.active
    ws1.title = "Roles"
    row = _add_excel_header(
        ws1, company,
        title_ar="تقرير صلاحيات الشركة — الأدوار",
        title_en="Company Permissions — Roles",
        cols_count=3,
    )

    headers = ["الدور / Role", "عدد الصلاحيات / Perm Count", "عدد المستخدمين / Users"]
    for i, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=i, value=h)
        _style_excel_header(cell)
    ws1.row_dimensions[row].height = 26
    row += 1

    roles = CustomRole.objects.filter(company=company).order_by('name')
    for idx, r in enumerate(roles):
        is_even = idx % 2 == 0
        users_count = UserRole.objects.filter(role=r).count()
        vals = [r.name, r.permissions.count(), users_count]
        for i, v in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=i, value=v)
            _style_excel_data(cell, is_even)
        row += 1

    for col, width in [('A', 35), ('B', 20), ('C', 20)]:
        ws1.column_dimensions[col].width = width

    row += 2
    _add_excel_footer(ws1, row, cols_count=3)

    # ─── Sheet 2: All Users ────────────────
    ws2 = wb.create_sheet("Users")
    row2 = _add_excel_header(
        ws2, company,
        title_ar="تقرير صلاحيات الشركة — المستخدمين",
        title_en="Company Permissions — Users",
        cols_count=4,
    )

    headers = ["المستخدم / User", "اسم المستخدم / Username", "الأدوار / Roles", "الاستثناءات / Overrides"]
    for i, h in enumerate(headers, 1):
        cell = ws2.cell(row=row2, column=i, value=h)
        _style_excel_header(cell)
    ws2.row_dimensions[row2].height = 26
    row2 += 1

    from django.contrib.auth import get_user_model
    User = get_user_model()
    users = User.objects.filter(company=company).order_by('username')
    for idx, u in enumerate(users):
        is_even = idx % 2 == 0
        name = f"{u.first_name} {u.last_name}".strip() or u.username
        roles_list = ", ".join([ur.role.name for ur in UserRole.objects.filter(user=u).select_related('role')]) or "—"
        overrides_count = UserPermissionOverride.objects.filter(user=u).count()
        vals = [name, u.username, roles_list, overrides_count]
        for i, v in enumerate(vals, 1):
            cell = ws2.cell(row=row2, column=i, value=v)
            _style_excel_data(cell, is_even)
        row2 += 1

    for col, width in [('A', 30), ('B', 20), ('C', 40), ('D', 15)]:
        ws2.column_dimensions[col].width = width

    row2 += 2
    _add_excel_footer(ws2, row2, cols_count=4)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = f'attachment; filename="company_{company.id}_permissions.xlsx"'
    return resp


# ══════════════════════════════════════
# ROLE EXPORT — PDF
# ══════════════════════════════════════
def export_role_pdf(role: CustomRole) -> HttpResponse:
    styles = _pdf_styles()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=25*mm,
    )

    elements = _build_pdf_header(
        role.company,
        subtitle_ar=f"صلاحيات الدور: {role.name}",
        subtitle_en=f"Role Permissions: {role.name}",
    )

    # Table
    headers = [
        ar_text("الصلاحية"),
        "Permission",
        ar_text("النطاق"),
        "Scope",
    ]

    data = [headers]
    perms = role.permissions.all().order_by('permission')
    for p in perms:
        data.append([
            ar_text(perm_label(p.permission)),
            perm_label_en(p.permission),
            ar_text(scope_label(p.scope)),
            scope_label_en(p.scope),
        ])

    table = Table(data, colWidths=[50*mm, 50*mm, 40*mm, 40*mm], repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD_NAME),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        # Body
        ('FONTNAME', (0, 1), (-1, -1), PDF_FONT_NAME),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
        ('GRID', (0, 0), (-1, -1), 0.3, BORDER_GRAY),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements, onFirstPage=_add_page_footer, onLaterPages=_add_page_footer)

    resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    resp['Content-Disposition'] = f'attachment; filename="role_{role.id}_permissions.pdf"'
    return resp


# ══════════════════════════════════════
# USER EXPORT — PDF
# ══════════════════════════════════════
def export_user_pdf(user) -> HttpResponse:
    styles = _pdf_styles()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=25*mm,
    )

    user_name = f"{user.first_name} {user.last_name}".strip() or user.username

    elements = _build_pdf_header(
        user.company,
        subtitle_ar=f"صلاحيات المستخدم: {user_name}",
        subtitle_en=f"User Permissions: {user_name}",
    )

    # User info box
    info_data = [
        [ar_text("اسم المستخدم"), user.username],
        [ar_text("البريد الإلكتروني"), user.email or "—"],
    ]
    info_table = Table(info_data, colWidths=[50*mm, 120*mm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), LIGHT_GRAY),
        ('FONTNAME', (0, 0), (0, -1), PDF_FONT_BOLD_NAME),
        ('FONTNAME', (1, 0), (1, -1), PDF_FONT_NAME),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.3, BORDER_GRAY),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10))

    # Roles section
    elements.append(Paragraph(ar_text("الأدوار المعينة / Assigned Roles"), styles["heading"]))

    user_roles = UserRole.objects.filter(user=user).select_related('role')
    if user_roles.exists():
        role_data = [[ar_text("الدور"), "Role"]]
        for ur in user_roles:
            role_data.append([ar_text(ur.role.name), ur.role.name])
        role_table = Table(role_data, colWidths=[85*mm, 85*mm], repeatRows=1)
        role_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), CYAN),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD_NAME),
            ('FONTNAME', (0, 1), (-1, -1), PDF_FONT_NAME),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.3, BORDER_GRAY),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(role_table)
    else:
        elements.append(Paragraph(ar_text("— لا توجد أدوار / No roles —"), styles["body"]))

    elements.append(Spacer(1, 12))

    # Overrides section
    elements.append(Paragraph(ar_text("الاستثناءات الشخصية / Personal Overrides"), styles["heading"]))

    overrides = UserPermissionOverride.objects.filter(user=user).order_by('permission')
    if overrides.exists():
        ov_data = [[
            ar_text("الصلاحية"), "Permission",
            ar_text("النطاق"), "Scope",
            ar_text("الحالة"), "Status",
        ]]
        for o in overrides:
            status_ar = "مسموح" if o.is_granted else "ممنوع"
            status_en = "Granted" if o.is_granted else "Denied"
            ov_data.append([
                ar_text(perm_label(o.permission)),
                perm_label_en(o.permission),
                ar_text(scope_label(o.scope)),
                scope_label_en(o.scope),
                ar_text(status_ar),
                status_en,
            ])
        ov_table = Table(ov_data, colWidths=[35*mm, 35*mm, 25*mm, 25*mm, 25*mm, 25*mm], repeatRows=1)
        ov_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), CYAN),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD_NAME),
            ('FONTNAME', (0, 1), (-1, -1), PDF_FONT_NAME),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.3, BORDER_GRAY),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(ov_table)
    else:
        elements.append(Paragraph(ar_text("— لا توجد استثناءات / No overrides —"), styles["body"]))

    doc.build(elements, onFirstPage=_add_page_footer, onLaterPages=_add_page_footer)

    resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    resp['Content-Disposition'] = f'attachment; filename="user_{user.id}_permissions.pdf"'
    return resp


# ══════════════════════════════════════
# COMPANY EXPORT — PDF
# ══════════════════════════════════════
def export_company_pdf(company) -> HttpResponse:
    styles = _pdf_styles()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=25*mm,
    )

    elements = _build_pdf_header(
        company,
        subtitle_ar="تقرير صلاحيات الشركة",
        subtitle_en="Company Permissions Report",
    )

    # Roles table
    elements.append(Paragraph(ar_text("الأدوار المخصصة / Custom Roles"), styles["heading"]))

    roles = CustomRole.objects.filter(company=company).order_by("name")
    if roles.exists():
        role_data = [[
            ar_text("الدور"), "Role",
            ar_text("عدد الصلاحيات"), "Perms",
            ar_text("عدد المستخدمين"), "Users",
        ]]
        for r in roles:
            users_count = UserRole.objects.filter(role=r).count()
            role_data.append([
                ar_text(r.name), r.name,
                str(r.permissions.count()), str(r.permissions.count()),
                str(users_count), str(users_count),
            ])
        role_table = Table(role_data, colWidths=[35*mm, 35*mm, 25*mm, 25*mm, 25*mm, 25*mm], repeatRows=1)
        role_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD_NAME),
            ("FONTNAME", (0, 1), (-1, -1), PDF_FONT_NAME),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.3, BORDER_GRAY),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(role_table)
    else:
        elements.append(Paragraph(ar_text("— لا توجد أدوار مخصصة —"), styles["body"]))

    elements.append(Spacer(1, 15))

    # Users table
    elements.append(Paragraph(ar_text("المستخدمين والأدوار / Users & Roles"), styles["heading"]))

    from django.contrib.auth import get_user_model
    User = get_user_model()
    users = User.objects.filter(company=company).order_by("username")

    if users.exists():
        user_data = [[
            ar_text("الاسم"), "Name",
            ar_text("اسم المستخدم"), "Username",
            ar_text("الأدوار"), "Roles",
        ]]
        for u in users:
            name = f"{u.first_name} {u.last_name}".strip() or u.username
            roles_list = ", ".join([ur.role.name for ur in UserRole.objects.filter(user=u).select_related("role")]) or "—"
            user_data.append([
                ar_text(name), name,
                u.username, u.username,
                ar_text(roles_list), roles_list,
            ])
        user_table = Table(user_data, colWidths=[30*mm, 30*mm, 25*mm, 25*mm, 30*mm, 30*mm], repeatRows=1)
        user_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD_NAME),
            ("FONTNAME", (0, 1), (-1, -1), PDF_FONT_NAME),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.3, BORDER_GRAY),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(user_table)
    else:
        elements.append(Paragraph(ar_text("— لا يوجد مستخدمين —"), styles["body"]))

    doc.build(elements, onFirstPage=_add_page_footer, onLaterPages=_add_page_footer)

    resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    resp['Content-Disposition'] = f'attachment; filename="company_{company.id}_permissions.pdf"'
    return resp



